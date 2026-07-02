"""
file SAP — STAGE 1: CLEAN ALL ADDRESSES
=======================================
Port engine ไฟล์แรก (postal DB + n-gram matcher) มาทำความสะอาดที่อยู่ทุก location.
แหล่งหลัก = Local_Address__c (ไทยสะอาด) → extract ตำบล/อำเภอ/จังหวัด/ถนน
แล้ว validate/standardize กับ postal DB; component เป็น fallback; แถว default ไม่เชื่อ Local.

Inputs : Location_TH_cleanup_new.xlsx + thai-postal-codes_V2.xlsx
Output : SAP_Location_Stage1_AddressClean.xlsx (Locations_AddrClean | Summary | Notes)
Pure stdlib + pandas + openpyxl.
"""
import re, math
from pathlib import Path
from collections import Counter, defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

DB   = Path("/Users/phachon/Documents/DKSH/location-clean-up/input/thai-postal-codes_V2.xlsx")
DATA = Path("/Users/phachon/Documents/DKSH/location-clean-up/input/Location_TH_cleanup_new.xlsx")
OUT  = Path("/Users/phachon/Documents/DKSH/location-clean-up/output/SAP_Location_Stage1_AddressClean.xlsx")

ACCEPT, EXACT = 0.60, 0.999

for _label, _p in [("DB (postal database)", DB), ("DATA (file SAP)", DATA)]:
    if not _p.exists():
        raise FileNotFoundError(f"ไม่พบไฟล์ {_label}: {_p}\n  → แก้ path ตัวแปรนี้ที่ด้านบนของสคริปต์ให้ตรงกับตำแหน่งไฟล์จริงในเครื่อง")
OUT.parent.mkdir(parents=True, exist_ok=True)

# ── helpers (ported) ───────────────────────────────────────────────────────────
def clean_ws(s):
    if pd.isna(s): return None
    s = re.sub(r"\s{2,}", " ", str(s).strip()).strip(" ,")
    return s or None
TH_RE = re.compile(r"[\u0E00-\u0E7F]")
def script_of(s):
    if not s: return None
    t = TH_RE.search(s); l = re.search(r"[A-Za-z]", s)
    if t and not l: return "TH"
    if l and not t: return "EN"
    if t and l: return "TH" if len(TH_RE.findall(s)) >= len(re.findall(r"[A-Za-z]", s)) else "EN"
    return None
PFX = re.compile(r"^(ตำบล|แขวง|ต\.|อำเภอ|เขต|อ\.|จังหวัด|จ\.|tambon|khwaeng|amphoe|khet|"
                 r"sub-?district|district|province)\s*", re.I)
def norm_match(s):
    s = clean_ws(s)
    if not s: return None
    prev = None
    while prev != s:
        prev = s; s = PFX.sub("", s).strip()
    if re.search(r"[A-Za-z]", s): return re.sub(r"[^a-z0-9]", "", s.lower())
    return s
def norm_road(s):
    s = clean_ws(s)
    if not s: return None
    s = re.sub(r"^ถ\.\s*", "ถนน", s); s = re.sub(r"^ซ\.\s*", "ซอย", s)
    return s
def norm_soi(s):
    s = clean_ws(s)
    if not s: return None
    s = re.sub(r"ซ\.", "ซอย", s); s = re.sub(r"^ซ(?=\s)", "ซอย", s)
    s = re.sub(r"^soi\b", "Soi", s, flags=re.I)
    return re.sub(r"\s{2,}", " ", s).strip()

class NGramCosine:
    def __init__(self, ns=(2,3)): self.ns=ns; self.idf={}; self.dflt=1.0; self.vecs={}
    def _grams(self,s):
        s=f"  {s}  "; out=[]
        for n in self.ns: out+=[s[i:i+n] for i in range(len(s)-n+1)]
        return out
    def fit(self,names):
        names={n for n in names if n}; dfc=Counter()
        for nm in names:
            for g in set(self._grams(nm)): dfc[g]+=1
        N=len(names); self.idf={g:math.log((N+1)/(d+1))+1 for g,d in dfc.items()}; self.dflt=math.log(N+1)+1
        for nm in names: self.vecs[nm]=self._vec(nm)
    def _vec(self,s):
        c=Counter(self._grams(s)); v={g:cnt*self.idf.get(g,self.dflt) for g,cnt in c.items()}
        nrm=math.sqrt(sum(w*w for w in v.values())) or 1.0
        return {g:w/nrm for g,w in v.items()}
    def _sim(self,qv,name):
        cv=self.vecs.get(name) or self._vec(name)
        a,b=(qv,cv) if len(qv)<len(cv) else (cv,qv)
        return sum(w*b.get(g,0.0) for g,w in a.items())
    def best(self,query,cands):
        q=norm_match(query)
        if not q or not cands: return None,0.0
        qv=self._vec(q); s=max(((self._sim(qv,c),c) for c in cands),default=(0.0,None))
        return s[1],round(s[0],3)

# ── load DB + indices (ported) ─────────────────────────────────────────────────
print("Loading Postal DB …")
xl=pd.ExcelFile(DB); post=xl.parse("รหัสไปรษณีย์",dtype=str); tam=xl.parse("TambonDatabase",dtype=str)
tam_ix=tam.set_index("TambonID")
def rec(tid):
    r=tam_ix.loc[tid]; pid=r.ProvinceID; bkk=pid=="10"
    return dict(tid=tid,did=r.DistrictID,pid=pid,
                sub_th=("แขวง" if bkk else "ตำบล")+r.TambonThaiShort,sub_sh=r.TambonThaiShort,
                sub_en=r.TambonEngShort,dist_th=r.DistrictThai,dist_sh=r.DistrictThaiShort,
                dist_en=r.DistrictEng,dist_ensh=r.DistrictEngShort,prov_th=r.ProvinceThai,prov_en=r.ProvinceEng,bkk=bkk)
postal_tambons=defaultdict(list); postal_prov=defaultdict(Counter)
for _,row in post.iterrows():
    pc,tid=clean_ws(row.PostCode),row.TambonID
    if pc and tid in tam_ix.index:
        rc=rec(tid); postal_tambons[pc].append(rc); postal_prov[pc][rc["prov_th"]]+=1
PROV_TH=set(tam["ProvinceThai"].dropna())
PROV_EN_NORM={e.lower():t for e,t in zip(tam["ProvinceEng"],tam["ProvinceThai"]) if pd.notna(e)}
PROV_LIST=list(PROV_TH); BKK_ALIAS={"กรุงเทพ","กรุงเทพฯ","กทม","กทม.","bangkok","krungthep","bkk"}
PROV_ALIAS={"ayutthaya":"พระนครศรีอยุธยา","korat":"นครราชสีมา","sriracha":"ชลบุรี"}
import difflib
def resolve_prov(s):
    s=clean_ws(s)
    if not s: return None
    b=re.sub(r"\s*\bprovince\b\s*","",s,flags=re.I); b=re.sub(r"^(จังหวัด|จ\.|จ\s)\s*","",b).strip()
    low=re.sub(r"\s+","",b.lower()).replace("ฯ","")
    if low in BKK_ALIAS or b in {"กรุงเทพ","กรุงเทพฯ"}: return "กรุงเทพมหานคร"
    if low in PROV_ALIAS: return PROV_ALIAS[low]
    if b in PROV_TH: return b
    if low in PROV_EN_NORM: return PROV_EN_NORM[low]
    if TH_RE.search(b):
        m=difflib.get_close_matches(b,PROV_LIST,n=1,cutoff=0.82)
        if m: return m[0]
    else:
        m=difflib.get_close_matches(low,list(PROV_EN_NORM),n=1,cutoff=0.82)
        if m: return PROV_EN_NORM[m[0]]
    return None
prov_tambons=defaultdict(list); prov_districts=defaultdict(dict)
for tid in tam_ix.index:
    rc=rec(tid); prov_tambons[rc["prov_th"]].append(rc); prov_districts[rc["prov_th"]].setdefault(rc["did"],rc)
_alln=set()
for _c in ["TambonThaiShort","TambonEngShort","DistrictThaiShort","DistrictEngShort"]:
    for _v in tam[_c].dropna():
        _nm=norm_match(_v)
        if _nm: _alln.add(_nm)
M=NGramCosine(); M.fit(_alln)
print(f"  {len(postal_tambons)} postal codes | matcher vocab {len(M.idf):,}")

def best_rec(query,recs,level,sc):
    if not query or not recs: return None,0.0
    key=({"sub":"sub_sh","dist":"dist_sh"} if sc=="TH" else {"sub":"sub_en","dist":"dist_ensh"})[level]
    n2r={}
    for r in recs:
        nm=norm_match(r[key])
        if nm: n2r.setdefault(nm,r)
    name,score=M.best(query,list(n2r))
    return (n2r.get(name),score) if name else (None,0.0)
def best_pool(fields,recs,level,sc):
    br,bs,bsrc=None,0.0,None
    for fn,val in fields:
        sv=script_of(val) or sc
        r,s=best_rec(val,recs,level,sv)
        if s>bs: br,bs,bsrc=r,s,fn
    return br,bs,bsrc

# ── NEW: parse Local_Address__c (clean Thai) ───────────────────────────────────
def parse_local(s):
    s=clean_ws(s)
    if not s: return (None,None,None,None,None)
    mz=re.findall(r"\b(\d{5})\b",s); postal=mz[-1] if mz else None
    sub=re.search(r"(?:ตำบล|แขวง)\s*([^\s\d]+)",s)
    dis=re.search(r"(?:อำเภอ|เขต)\s*([^\s\d]+)",s)
    prov=re.search(r"จังหวัด\s*([^\s\d]+)",s)
    prov_v=prov.group(1) if prov else ("กรุงเทพมหานคร" if "กรุงเทพมหานคร" in s else None)
    cut=re.search(r"(ตำบล|แขวง|อำเภอ|เขต|เมืองพัทยา)",s)
    street=clean_ws(s[:cut.start()]) if cut else None
    return (sub.group(1) if sub else None, dis.group(1) if dis else None, prov_v, street, postal)
def parse_street_text(s):
    s=clean_ws(s)
    if not s: return ("","","","")
    house=moo=soi=road=""
    m=re.match(r"^(?:เลขที่\s*)?(\d+[\d/\-]*)",s); rest=s
    if m: house=m.group(1); rest=s[m.end():]
    mm=re.search(r"(?:หมู่ที่|หมู่|ม\.)\s*(\d+)",rest)
    if mm: moo=mm.group(1)
    ms=re.search(r"(ซอย\S+|ซ\.\S+)",rest)
    if ms: soi=norm_soi(ms.group(1)) or ""
    mr=re.search(r"(ถนน\S+|ถ\.\S+)",rest)
    if mr: road=norm_road(mr.group(1)) or ""
    return (house,moo,soi,road)

def disp_prov(p): return "กรุงเทพมหานคร" if p=="กรุงเทพมหานคร" else (f"จังหวัด{p}" if p else "")

HOUSE_LEAD = re.compile(r"^\d+[\d/\-]*(?=\s|$)")
def ensure_house_prefix(s):
    """เติม 'เลขที่' นำหน้าให้ข้อความที่ขึ้นต้นด้วยเลขบ้านแต่ยังไม่มีคำนี้ — ไม่แตะรายละเอียดอื่นในข้อความเลย."""
    s = clean_ws(s)
    if not s: return s
    if s.startswith("เลขที่"): return s
    if HOUSE_LEAD.match(s): return "เลขที่ " + s
    return s

# ── load file SAP + default detection ──────────────────────────────────────────
print("Loading file SAP …")
sap=pd.read_excel(DATA,sheet_name="Location clean up",dtype=str)
ac=sap['Address__c'].value_counts(); DEFAULT_SET=set(ac[ac>20].index)
print(f"  {len(sap)} locations | default Address__c values (>20x): {len(DEFAULT_SET)}")

print("Cleaning addresses …")
rows=[]
for i in sap.index:
    zip_raw=clean_ws(sap.at[i,'SVMXC__Zip__c']); local=sap.at[i,'Local_Address__c']
    addr_en=sap.at[i,'Address__c']; street_blob=sap.at[i,'SVMXC__Street__c']
    dist_c=clean_ws(sap.at[i,'District__c']); city_c=clean_ws(sap.at[i,'SVMXC__City__c']); state_c=clean_ws(sap.at[i,'SVMXC__State__c'])
    flags=[]; is_def=(addr_en in DEFAULT_SET)
    if is_def: flags.append("default_address_suspect")

    zip_postal=None
    if zip_raw:
        mz=re.search(r"\d{5}",zip_raw); zip_postal=mz.group() if mz else None
    sub_L=dist_L=prov_L=street_L=postal_L=None
    if not is_def: sub_L,dist_L,prov_L,street_L,postal_L=parse_local(local)
    postal = postal_L if (not is_def and postal_L) else zip_postal
    if zip_postal and postal_L and zip_postal!=postal_L: flags.append(f"postal_mismatch(zip={zip_postal}/local={postal_L})")

    covered=postal in postal_tambons
    if covered:
        prov_th=postal_prov[postal].most_common(1)[0][0]
        cands=postal_tambons[postal]; dist_recs=list({r['did']:r for r in cands}.values())
    else:
        prov_th=resolve_prov(prov_L or state_c or city_c)
        flags.append("postal_not_in_db" if postal else "no_postal")
        cands=prov_tambons.get(prov_th,[]) if prov_th else []
        dist_recs=list(prov_districts.get(prov_th,{}).values()) if prov_th else []

    sub_rec,sub_s,sub_src=best_pool([("local",sub_L),("District__c",dist_c),("City__c",city_c)],cands,"sub","TH")
    if sub_rec and sub_s>=ACCEPT:
        if sub_src!="local" or sub_s<0.95: flags.append(f"subdistrict:{sub_src}({sub_s:.2f})")
    else:
        if any([sub_L,dist_c]): flags.append("subdistrict:unmatched")
        sub_rec=None

    if sub_rec:
        dist_rec=next((r for r in dist_recs if r['did']==sub_rec['did']),sub_rec)
    elif len(dist_recs)==1:
        dist_rec=dist_recs[0]
    else:
        dr,ds,_=best_pool([("local",dist_L),("City__c",city_c),("State__c",state_c)],dist_recs,"dist","TH")
        dist_rec=dr if (dr and ds>=ACCEPT) else None

    house,moo,soi,road=parse_street_text(street_L or street_blob)
    source="local" if (not is_def and (sub_L or dist_L)) else "component"

    sub_th=sub_rec['sub_th'] if sub_rec else ""; sub_en=sub_rec['sub_en'] if sub_rec else ""
    dist_th=dist_rec['dist_th'] if dist_rec else ""; dist_en=dist_rec['dist_ensh'] if dist_rec else ""
    prov_native=disp_prov(prov_th); prov_en=(next((r['prov_en'] for r in cands),None) or (next((r['prov_en'] for r in prov_tambons.get(prov_th,[])),prov_th) if prov_th else "")) or ""
    tid=sub_rec['tid'] if sub_rec else ""; did=dist_rec['did'] if dist_rec else ""
    pid=(cands[0]['pid'] if cands else (sub_rec['pid'] if sub_rec else ""))
    house_col=f"เลขที่ {house}" if house else ""; moo_col=f"หมู่ที่ {moo}" if moo else ""
    # street portion: ใช้ Local ทั้งก้อน (สะอาด) ถ้ามี ไม่งั้นประกอบจาก blob — แต่บังคับให้มี 'เลขที่' นำหน้าเสมอ
    street_part = ensure_house_prefix(street_L) if (source=="local" and street_L) else " ".join([x for x in [house_col,moo_col,soi,road] if x])

    seg=[x for x in [street_part,sub_th,dist_th,prov_native,postal] if x]
    full_th=" ".join(seg)
    te=[x for x in [sub_en,dist_en,prov_en,postal] if x]
    full_en=", ".join(([house] if house else [])+te+["Thailand"])

    # status
    if not covered or not postal: status="NEEDS_REVIEW"
    elif not sub_rec and not dist_rec: status="NEEDS_REVIEW"; flags.append("incomplete")
    elif any("unmatched" in f for f in flags): status="AUTO_FIXED"
    elif sub_rec and sub_s<0.95 and sub_s>0: status="FUZZY_FIXED"
    elif sub_rec or dist_rec: status="AUTO_FIXED" if (source=="component" or not sub_rec or sub_s<EXACT) else "VERIFIED"
    else: status="NEEDS_REVIEW"
    if is_def and status=="VERIFIED": status="AUTO_FIXED"
    prov_s=1.0 if covered else (0.5 if prov_th else 0.0)
    sub_score=sub_s if sub_rec else (1.0 if not any([sub_L,dist_c]) else 0.0)
    dist_score=1.0 if (dist_rec and (len(dist_recs)==1 or sub_rec)) else (0.6 if dist_rec else 0.0)
    conf=round(100*(0.34*prov_s+0.33*min(dist_score,1)+0.33*min(sub_score,1)))

    rows.append(dict(house=house_col,moo=moo_col,soi=soi or "",road=road or "",street=street_part,
        sub=sub_th,sub_en=sub_en,dist=dist_th,dist_en=dist_en,prov=prov_native,prov_en=prov_en,
        postal=postal or "",tid=tid,did=did,pid=pid,full_th=full_th,full_en=full_en,
        status=status,conf=conf,source=source,
        default="Yes" if is_def else "",mismatch="Yes" if any("postal_mismatch" in f for f in flags) else "",
        flags="; ".join(flags) if flags else "ok"))

res=pd.DataFrame(rows,index=sap.index)

# ── output ─────────────────────────────────────────────────────────────────────
print("Writing workbook …")
out=sap.copy()
addmap={"sap_addr_house":"house","sap_addr_moo":"moo","sap_addr_soi":"soi","sap_addr_road":"road",
    "sap_addr_street":"street",
    "sap_addr_subdistrict":"sub","sap_addr_district":"dist","sap_addr_province":"prov",
    "sap_addr_subdistrict_en":"sub_en","sap_addr_district_en":"dist_en","sap_addr_province_en":"prov_en",
    "sap_addr_postal":"postal","sap_tambon_id":"tid","sap_district_id":"did","sap_province_id":"pid",
    "sap_addr_full_th":"full_th","sap_addr_full_en":"full_en","sap_addr_status":"status",
    "sap_addr_confidence":"conf","sap_addr_source":"source","default_address_suspect":"default",
    "postal_mismatch":"mismatch","sap_addr_flags":"flags"}
for col,src in addmap.items(): out[col]=res[src].values
ADDED=list(addmap.keys())

FILL={"VERIFIED":PatternFill("solid",start_color="C8E6C9"),"AUTO_FIXED":PatternFill("solid",start_color="BBDEFB"),
      "FUZZY_FIXED":PatternFill("solid",start_color="FFF9C4"),"NEEDS_REVIEW":PatternFill("solid",start_color="FFCDD2")}
wb=Workbook(); del wb["Sheet"]; ws=wb.create_sheet("Locations_AddrClean")
cols=list(out.columns); addset={cols.index(c)+1 for c in ADDED}
scol=cols.index("sap_addr_status")+1; fcol=cols.index("sap_addr_full_th")+1
data=[[v if pd.notna(v) else None for v in r] for r in out.itertuples(index=False)]
allrows=[cols]+data
for ri,row in enumerate(allrows,1):
    for ci,val in enumerate(row,1):
        c=ws.cell(ri,ci,val)
        if ri==1:
            c.font=Font(bold=True,color="FFFFFF",name="Arial",size=10)
            c.fill=PatternFill("solid",start_color="F57C00") if ci in addset else PatternFill("solid",start_color="37474F")
            c.alignment=Alignment(horizontal="center",wrap_text=True)
        else:
            c.font=Font(name="Arial",size=10)
            if ci==scol and val in FILL: c.fill=FILL[val]
            elif ci==fcol and row[scol-1] in FILL: c.fill=FILL[row[scol-1]]
ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
for ci in range(1,len(cols)+1):
    w=max((len(str(r[ci-1] or "")) for r in allrows[:60]),default=10)
    ws.column_dimensions[ws.cell(1,ci).column_letter].width=min(w+2,46)

st=Counter(res['status']); N=len(res)
ws2=wb.create_sheet("Summary")
rowsS=[["file SAP — Stage 1 Address Clean — Summary",""],["",""],["สถานะ","จำนวน","%"]]
for k in ["VERIFIED","AUTO_FIXED","FUZZY_FIXED","NEEDS_REVIEW"]:
    rowsS.append([k,st.get(k,0),f"{100*st.get(k,0)/N:.1f}%"])
rowsS+=[["รวม",N,"100%"],["",""],
    ["default_address_suspect",int((res['default']=="Yes").sum()),"ที่อยู่น่าจะเป็น HQ default"],
    ["postal_mismatch",int((res['mismatch']=="Yes").sum()),"zip ≠ postal ใน local"],
    ["source=local",int((res['source']=="local").sum()),"ใช้ Local_Address__c"],
    ["source=component",int((res['source']=="component").sum()),"ใช้ field components"]]
for ri,row in enumerate(rowsS,1):
    for ci,val in enumerate(row,1):
        c=ws2.cell(ri,ci,val)
        if ri==1: c.font=Font(bold=True,size=13)
        if ci==1 and val in FILL: c.fill=FILL[val]
ws2.column_dimensions['A'].width=26; ws2.column_dimensions['B'].width=12; ws2.column_dimensions['C'].width=34

ws3=wb.create_sheet("Notes")
def put(r,t,b=False):
    c=ws3.cell(r,1,t); c.font=Font(bold=b,name="Arial",size=12 if (b and r==1) else (11 if b else 10)); c.alignment=Alignment(wrap_text=True,vertical="top")
r=1
for t,b in [("file SAP — Stage 1: Clean Address — วิธีการ",True),("",False),
 ("• Port engine ไฟล์แรก: postal DB + n-gram cosine matcher (offline)",False),
 ("• แหล่งหลัก = Local_Address__c (ไทยสะอาด) → extract ตำบล/อำเภอ/จังหวัด/ถนน แล้ว validate กับ postal DB",False),
 ("• component (District__c/City__c/State__c) = fallback; แถว default (Address__c ซ้ำ >20x) ไม่เชื่อ Local",False),
 ("• จังหวัด = postal (authoritative); ตำบล/อำเภอ = matcher; ที่อยู่ออกเป็นทางการ TH+EN + official IDs",False),
 ("",False),("คอลัมน์ที่เพิ่ม (sap_addr_*)",True),
 ("sap_addr_house/moo/soi/road — รายละเอียด (เต็มรูปแบบ)",False),
 ("sap_addr_subdistrict/district/province (+_en) — ทางการ + sap_addr_postal",False),
 ("sap_tambon_id/district_id/province_id — รหัสทางการ (ไว้ join)",False),
 ("sap_addr_full_th / full_en — ที่อยู่ประกอบเต็ม 2 ภาษา",False),
 ("sap_addr_status (สี) / confidence / source (local|component) / flags",False),
 ("default_address_suspect / postal_mismatch — ธงเตือน",False),
 ("",False),("สถานะสี: เขียว VERIFIED · ฟ้า AUTO_FIXED · เหลือง FUZZY · แดง NEEDS_REVIEW",True),
 ("",False),("ขั้นต่อไป: Stage 2 = multi-signal matching (ใช้ที่อยู่ clean นี้เป็น input)",False)]:
    put(r,t,b); r+=1
ws3.column_dimensions['A'].width=100
wb.save(OUT)

print("\n"+"="*58+"\n  STAGE 1 COMPLETE")
print(f"  Locations: {N:,}")
for k in ["VERIFIED","AUTO_FIXED","FUZZY_FIXED","NEEDS_REVIEW"]:
    print(f"    {k:13} {st.get(k,0):6,} ({100*st.get(k,0)/N:4.1f}%)")
print(f"  default {int((res['default']=='Yes').sum()):,} | source=local {int((res['source']=='local').sum()):,}")
print(f"  Output: {OUT.name}\n"+"="*58)
for s in ["VERIFIED","AUTO_FIXED","NEEDS_REVIEW"]:
    ex=res[res['status']==s].head(1)
    if len(ex): print(f"  [{s}] {ex.iloc[0]['full_th'][:84]}")