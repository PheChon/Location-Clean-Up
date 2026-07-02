"""
SAP <-> MASTER — STAGE 2: MULTI-SIGNAL MATCHING (non-maskey -> master)
========================================================================
จับคู่ location ที่ไม่มี "maskey" (SAP code ไม่ตรง master หรือไม่มี code) เข้ากับ master
ด้วย 3 สัญญาณ: M1 ชื่อ + M2 Organization sibling + M4 ที่อยู่ (เทียบ ID ทางการ ไม่ fuzzy text)
ไม่ใช้ dedup key เดิมของ ServiceMax (Location Duplication Check/Clean street/Suspect) ตามที่ตกลง
ไม่ทำ duplicate-clustering ระหว่าง non-maskey กันเอง (เลื่อนไป stage หลัง) — pass นี้จับ non-maskey vs master เท่านั้น

SELF-CONTAINED: คำนวณ maskey-join (Step 1 logic) + address-clean (Stage 1 engine) ซ้ำจาก 3 ไฟล์ต้นทาง
ไม่อ่าน output ไฟล์ก่อนหน้า กันไฟล์เก่าเพี้ยนถ้า logic เคยแก้ไปแล้ว

Inputs : Location_TH_cleanup_new.xlsx + TH44_Addresses_Cleaned_V2.xlsx + thai-postal-codes_V2.xlsx
Output : SAP_Master_Stage2_Matched.xlsx (Locations_Matched | Summary | Notes) — ไฟล์เดียวครบทุกคอลัมน์
Pure stdlib + pandas + openpyxl.
"""
import re, math, difflib
from pathlib import Path
from collections import Counter, defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

DB   = Path("/Users/phachon/Documents/DKSH/location-clean-up/input/thai-postal-codes_V2.xlsx")
DATA = Path("/Users/phachon/Documents/DKSH/location-clean-up/input/Location_TH_cleanup_new.xlsx")
MASTER = Path("/Users/phachon/Documents/DKSH/location-clean-up/output/TH44_Addresses_Cleaned_V2.xlsx")
OUT  = Path("/Users/phachon/Documents/DKSH/location-clean-up/output/SAP_Location_Stage2-2_AddressClean.xlsx")

ACCEPT, EXACT   = 0.60, 0.999   # Stage-1 address-match thresholds
NAME_FLOOR      = 0.55          # ต่ำกว่านี้ไม่นับเป็น candidate เลย (ตาม Step1)
NAME_STRONG     = 0.85          # ชื่อแรงพอจะ AUTO ได้เองถ้าไม่ generic

for _label, _p in [("DB (postal database)", DB), ("DATA (file SAP)", DATA), ("MASTER (master file)", MASTER)]:
    if not _p.exists():
        raise FileNotFoundError(f"ไม่พบไฟล์ {_label}: {_p}\n  → แก้ path ตัวแปรนี้ที่ด้านบนของสคริปต์ให้ตรงกับตำแหน่งไฟล์จริงในเครื่อง")
OUT.parent.mkdir(parents=True, exist_ok=True)

# ══════════════════════════════════════════════════════════════════════════════
# PART A — ported Stage-1 address-cleaning engine (verbatim, bug-fixed version)
# ══════════════════════════════════════════════════════════════════════════════
def clean_ws(s):
    if pd.isna(s): return None
    s = re.sub(r"\s{2,}", " ", str(s).strip()).strip(" ,")
    return s or None
TH_RE = re.compile(r"[\u0E00-\u0E7F]")
def script_of(s):
    if not s: return None
    t = TH_RE.search(s); l = re.search(r"[A-Za-z]", s)
    if t and not l: return "TH"
    if l and not t: return "EN"
    if t and l: return "TH" if len(TH_RE.findall(s)) >= len(re.findall(r"[A-Za-z]", s)) else "EN"
    return None
PFX = re.compile(r"^(ตำบล|แขวง|ต\.|อำเภอ|เขต|อ\.|จังหวัด|จ\.|tambon|khwaeng|amphoe|khet|"
                 r"sub-?district|district|province)\s*", re.I)
def norm_match(s):
    s = clean_ws(s)
    if not s: return None
    prev = None
    while prev != s:
        prev = s; s = PFX.sub("", s).strip()
    if re.search(r"[A-Za-z]", s): return re.sub(r"[^a-z0-9]", "", s.lower())
    return s
def norm_road(s):
    s = clean_ws(s)
    if not s: return None
    s = re.sub(r"^ถ\.\s*", "ถนน", s); s = re.sub(r"^ซ\.\s*", "ซอย", s)
    return s
def norm_soi(s):
    s = clean_ws(s)
    if not s: return None
    s = re.sub(r"ซ\.", "ซอย", s); s = re.sub(r"^ซ(?=\s)", "ซอย", s)
    s = re.sub(r"^soi\b", "Soi", s, flags=re.I)
    return re.sub(r"\s{2,}", " ", s).strip()

class NGramCosine:
    def __init__(self, ns=(2,3)): self.ns=ns; self.idf={}; self.dflt=1.0; self.vecs={}
    def _grams(self,s):
        s=f"  {s}  "; out=[]
        for n in self.ns: out+=[s[i:i+n] for i in range(len(s)-n+1)]
        return out
    def fit(self,names):
        names={n for n in names if n}; dfc=Counter()
        for nm in names:
            for g in set(self._grams(nm)): dfc[g]+=1
        N=len(names); self.idf={g:math.log((N+1)/(d+1))+1 for g,d in dfc.items()}; self.dflt=math.log(N+1)+1
        for nm in names: self.vecs[nm]=self._vec(nm)
    def _vec(self,s):
        c=Counter(self._grams(s)); v={g:cnt*self.idf.get(g,self.dflt) for g,cnt in c.items()}
        nrm=math.sqrt(sum(w*w for w in v.values())) or 1.0
        return {g:w/nrm for g,w in v.items()}
    def _sim(self,qv,name):
        cv=self.vecs.get(name) or self._vec(name)
        a,b=(qv,cv) if len(qv)<len(cv) else (cv,qv)
        return sum(w*b.get(g,0.0) for g,w in a.items())
    def best(self,query,cands):
        q=norm_match(query)
        if not q or not cands: return None,0.0
        qv=self._vec(q); s=max(((self._sim(qv,c),c) for c in cands),default=(0.0,None))
        return s[1],round(s[0],3)

print("Loading Postal DB …")
xl=pd.ExcelFile(DB); post=xl.parse("รหัสไปรษณีย์",dtype=str); tam=xl.parse("TambonDatabase",dtype=str)
tam_ix=tam.set_index("TambonID")
def rec(tid):
    r=tam_ix.loc[tid]; pid=r.ProvinceID; bkk=pid=="10"
    return dict(tid=tid,did=r.DistrictID,pid=pid,
                sub_th=("แขวง" if bkk else "ตำบล")+r.TambonThaiShort,sub_sh=r.TambonThaiShort,
                sub_en=r.TambonEngShort,dist_th=r.DistrictThai,dist_sh=r.DistrictThaiShort,
                dist_en=r.DistrictEng,dist_ensh=r.DistrictEngShort,prov_th=r.ProvinceThai,prov_en=r.ProvinceEng,bkk=bkk)
postal_tambons=defaultdict(list); postal_prov=defaultdict(Counter)
for _,row in post.iterrows():
    pc,tid=clean_ws(row.PostCode),row.TambonID
    if pc and tid in tam_ix.index:
        rc=rec(tid); postal_tambons[pc].append(rc); postal_prov[pc][rc["prov_th"]]+=1
PROV_TH=set(tam["ProvinceThai"].dropna())
PROV_EN_NORM={e.lower():t for e,t in zip(tam["ProvinceEng"],tam["ProvinceThai"]) if pd.notna(e)}
PROV_LIST=list(PROV_TH); BKK_ALIAS={"กรุงเทพ","กรุงเทพฯ","กทม","กทม.","bangkok","krungthep","bkk"}
PROV_ALIAS={"ayutthaya":"พระนครศรีอยุธยา","korat":"นครราชสีมา","sriracha":"ชลบุรี"}
def resolve_prov(s):
    s=clean_ws(s)
    if not s: return None
    b=re.sub(r"\s*\bprovince\b\s*","",s,flags=re.I); b=re.sub(r"^(จังหวัด|จ\.|จ\s)\s*","",b).strip()
    low=re.sub(r"\s+","",b.lower()).replace("ฯ","")
    if low in BKK_ALIAS or b in {"กรุงเทพ","กรุงเทพฯ"}: return "กรุงเทพมหานคร"
    if low in PROV_ALIAS: return PROV_ALIAS[low]
    if b in PROV_TH: return b
    if low in PROV_EN_NORM: return PROV_EN_NORM[low]
    if TH_RE.search(b):
        m=difflib.get_close_matches(b,PROV_LIST,n=1,cutoff=0.82)
        if m: return m[0]
    else:
        m=difflib.get_close_matches(low,list(PROV_EN_NORM),n=1,cutoff=0.82)
        if m: return PROV_EN_NORM[m[0]]
    return None
prov_tambons=defaultdict(list); prov_districts=defaultdict(dict)
for tid in tam_ix.index:
    rc=rec(tid); prov_tambons[rc["prov_th"]].append(rc); prov_districts[rc["prov_th"]].setdefault(rc["did"],rc)
_alln=set()
for _c in ["TambonThaiShort","TambonEngShort","DistrictThaiShort","DistrictEngShort"]:
    for _v in tam[_c].dropna():
        _nm=norm_match(_v)
        if _nm: _alln.add(_nm)
M=NGramCosine(); M.fit(_alln)
print(f"  {len(postal_tambons)} postal codes | matcher vocab {len(M.idf):,}")

def best_rec(query,recs,level,sc):
    if not query or not recs: return None,0.0
    key=({"sub":"sub_sh","dist":"dist_sh"} if sc=="TH" else {"sub":"sub_en","dist":"dist_ensh"})[level]
    n2r={}
    for r in recs:
        nm=norm_match(r[key])
        if nm: n2r.setdefault(nm,r)
    name,score=M.best(query,list(n2r))
    return (n2r.get(name),score) if name else (None,0.0)
def best_pool(fields,recs,level,sc):
    br,bs,bsrc=None,0.0,None
    for fn,val in fields:
        sv=script_of(val) or sc
        r,s=best_rec(val,recs,level,sv)
        if s>bs: br,bs,bsrc=r,s,fn
    return br,bs,bsrc

def strip_company_prefix(s):
    s = re.sub(r"_x000D_", "\n", str(s))
    lines = [p.strip() for p in re.split(r"[\n\r]+", s) if p.strip()]
    while len(lines) > 1:
        ln = lines[0]
        if re.search(r"(co\.|ltd|limited|บริษัท|จำกัด|\binc\b|公司)", ln, re.I) and not re.search(r"\d{5}", ln):
            lines.pop(0)
        else:
            break
    return " ".join(lines)
def disp_prov(p): return "กรุงเทพมหานคร" if p=="กรุงเทพมหานคร" else (f"จังหวัด{p}" if p else "")
def clean_admin_token(s):
    s=clean_ws(s)
    if not s: return None
    s=re.sub(r"\s*(sub-?district|district)\s*,?\s*$","",s,flags=re.I)
    s=re.sub(r"^(จ\.|อ\.|ต\.)\s*","",s)
    return clean_ws(s.rstrip(", "))

def parse_component_street(s):
    s=clean_ws(strip_company_prefix(s)) if pd.notna(s) else None
    if not s: return ("","","","",None,None)
    house=moo=soi=road=""; sub_hint=dist_hint=None
    m=re.match(r"^(?:เลขที่\s*)?(\d+[\d/\-]*)",s)
    if m: house=m.group(1)
    mm=re.search(r"(?:หมู่ที่|หมู่|ม\.|moo)\s*\.?\s*(\d+)",s,re.I)
    if mm: moo=mm.group(1)
    _SOI_STOP=r"(?=\s+(?:ถนน|ถ\.|ตำบล|แขวง|อำเภอ|เขต|จังหวัด)|,|$)"
    _ROAD_STOP=r"(?=\s+(?:ซอย|ซ\.|ตำบล|แขวง|อำเภอ|เขต|จังหวัด)|,|$)"
    ms=re.search(r"((?:ซอย|ซ\.)\s*.*?)"+_SOI_STOP,s)
    if ms:
        _v=norm_soi(ms.group(1)) or ""
        if _v and _v!="ซอย": soi=_v
    mr=re.search(r"((?:ถนน|ถ\.)\s*.*?)"+_ROAD_STOP,s)
    if mr:
        _v=norm_road(mr.group(1)) or ""
        if _v and _v!="ถนน": road=_v
    sh=re.search(r"(?:ตำบล|แขวง)\s*([ก-๙]+)",s)
    if not sh: sh=re.search(r"([A-Za-z][A-Za-z ]*?)\s+Sub-?district",s,re.I)
    if sh: sub_hint=sh.group(1).strip()
    dh=re.search(r"(?:อำเภอ|เขต)\s*([ก-๙]+)",s)
    if not dh: dh=re.search(r"([A-Za-z][A-Za-z ]*?)\s+District",s,re.I)
    if dh: dist_hint=dh.group(1).strip()
    return (house,moo,soi,road,sub_hint,dist_hint)

# ══════════════════════════════════════════════════════════════════════════════
# PART B — ported Step-1 code/name normalization
# ══════════════════════════════════════════════════════════════════════════════
def canon(s):
    if pd.isna(s): return None
    s = str(s).strip()
    return (s.lstrip('0') or '0') if s else None
SUFFIX = re.compile(
    r'\b(co\.?\s*,?\s*ltd\.?|company\s*limited|limited|ltd\.?|public\s*company\s*limited|'
    r'plc|inc\.?|corporation|corp\.?|บริษัท|จำกัด|มหาชน|\(thailand\)|\(ประเทศไทย\))\b', re.I)
def norm_name(s):
    if pd.isna(s): return ''
    s = str(s).lower()
    s = re.sub(r'\(.*?\)', ' ', s)
    s = re.split(r'\s[-–]\s', s)[0]
    s = SUFFIX.sub(' ', s)
    s = re.sub(r'[^a-z0-9ก-๙]+', '', s)
    return s.strip()
def name_sim(a, b):
    if not a or not b: return None
    if a == b or a in b or b in a: return 1.0
    return difflib.SequenceMatcher(None, a, b).ratio()

print("Loading file SAP + master …")
sap = pd.read_excel(DATA, sheet_name="Location clean up", dtype=str)
mas = pd.read_excel(MASTER, sheet_name="Addresses", dtype=str)
sap['_k'] = sap['SVMX_SAP_Code__c'].apply(canon)
mas['_k'] = mas['Cuscode'].apply(canon)
print(f"  SAP {len(sap)} locations | master {len(mas)} customers")

mas_keys = set(mas['_k'].dropna())
sap['maskey_status'] = sap['_k'].apply(lambda k: 'MASKEY' if (pd.notna(k) and k in mas_keys)
                                        else ('ORPHAN_SAP' if pd.notna(k) else 'NO_SAP_CODE'))
mas_by_k = {r['_k']: r for _, r in mas.iterrows()}
print("  ", Counter(sap['maskey_status']))

# ══════════════════════════════════════════════════════════════════════════════
# PART C — Stage-1 logic: clean address for ALL 18,007 rows (component-based)
# ══════════════════════════════════════════════════════════════════════════════
print("Cleaning addresses (component-based, ported Stage 1) …")
addr_rows=[]
for i in sap.index:
    zip_raw=clean_ws(sap.at[i,'SVMXC__Zip__c']); street_blob=sap.at[i,'SVMXC__Street__c']
    dist_c=clean_admin_token(sap.at[i,'District__c'])
    city_c=clean_admin_token(sap.at[i,'SVMXC__City__c'])
    state_c=clean_admin_token(sap.at[i,'SVMXC__State__c'])
    flags=[]

    postal=None
    if zip_raw:
        mz=re.search(r"\d{5}",zip_raw); postal=mz.group() if mz else None
    covered=postal in postal_tambons
    if covered:
        prov_th=postal_prov[postal].most_common(1)[0][0]
        cands=postal_tambons[postal]; dist_recs=list({r['did']:r for r in cands}.values())
    else:
        prov_th=resolve_prov(state_c) or resolve_prov(city_c)
        flags.append("postal_not_in_db" if postal else "no_postal")
        cands=prov_tambons.get(prov_th,[]) if prov_th else []
        dist_recs=list(prov_districts.get(prov_th,{}).values()) if prov_th else []

    house,moo,soi,road,sub_hint,dist_hint=parse_component_street(street_blob)

    sc_prov=resolve_prov(state_c) or resolve_prov(city_c)
    prov_conflict=False
    if covered and sc_prov and sc_prov!=prov_th:
        flags.append(f"conflict:prov(state/city={sc_prov}!=postal={prov_th})"); prov_conflict=True

    dist_narrowed=None
    if len(dist_recs)>1:
        def _not_just_province(v):
            return bool(v) and resolve_prov(v)!=prov_th
        cand_fields=[(fn,v) for fn,v in [("City__c",city_c),("street_hint",dist_hint),("State__c",state_c)] if _not_just_province(v)]
        if cand_fields:
            dr0,ds0,dsrc0=best_pool(cand_fields,dist_recs,"dist","TH")
            if dr0 and ds0>=ACCEPT:
                dist_narrowed=dr0
                flags.append(f"district_narrowed:{dsrc0}({ds0:.2f})")
    sub_cands=[r for r in cands if r['did']==dist_narrowed['did']] if dist_narrowed else cands

    sub_rec,sub_s,sub_src=best_pool([("District__c",dist_c),("street_hint",sub_hint),("City__c",city_c)],sub_cands,"sub","TH")
    if sub_rec and sub_s>=ACCEPT:
        if sub_s<0.95: flags.append(f"subdistrict:{sub_src}({sub_s:.2f})")
    else:
        if any([dist_c,sub_hint]): flags.append("subdistrict:unmatched")
        sub_rec=None

    if sub_rec:
        dist_rec=next((r for r in dist_recs if r['did']==sub_rec['did']),sub_rec)
    elif dist_narrowed:
        dist_rec=dist_narrowed
    elif len(dist_recs)==1:
        dist_rec=dist_recs[0]
    else:
        dr,ds,_=best_pool([("City__c",city_c),("street_hint",dist_hint),("State__c",state_c)],dist_recs,"dist","TH")
        dist_rec=dr if (dr and ds>=ACCEPT) else None

    source="component" if (dist_c or city_c or sub_hint or dist_hint or road or house) else ("postal_only" if postal else "none")

    sub_th=sub_rec['sub_th'] if sub_rec else ""; sub_en=sub_rec['sub_en'] if sub_rec else ""
    dist_th=dist_rec['dist_th'] if dist_rec else ""; dist_en=dist_rec['dist_ensh'] if dist_rec else ""
    prov_native=disp_prov(prov_th); prov_en=(next((r['prov_en'] for r in cands),None) or (next((r['prov_en'] for r in prov_tambons.get(prov_th,[])),prov_th) if prov_th else "")) or ""
    tid=sub_rec['tid'] if sub_rec else ""; did=dist_rec['did'] if dist_rec else ""
    pid=(cands[0]['pid'] if cands else (sub_rec['pid'] if sub_rec else ""))
    house_col=f"เลขที่ {house}" if house else ""; moo_col=f"หมู่ที่ {moo}" if moo else ""
    street_part=" ".join([x for x in [house_col,moo_col,soi,road] if x])

    seg=[x for x in [street_part,sub_th,dist_th,prov_native,postal] if x]
    full_th=" ".join(seg)
    te=[x for x in [sub_en,dist_en,prov_en,postal] if x]
    full_en=", ".join(([house] if house else [])+te+["Thailand"]) if (house or te) else ""

    if not postal and not prov_th: status="NEEDS_REVIEW"; flags.append("no_location_data")
    elif not covered: status="NEEDS_REVIEW"
    elif prov_conflict: status="NEEDS_REVIEW"
    elif not sub_rec and not dist_rec: status="NEEDS_REVIEW"; flags.append("incomplete")
    elif not sub_rec: status="AUTO_FIXED"
    elif sub_s<0.95 and sub_s>0: status="FUZZY_FIXED"
    elif sub_s>=EXACT: status="VERIFIED"
    else: status="AUTO_FIXED"
    prov_s=1.0 if covered else (0.5 if prov_th else 0.0)
    sub_score=sub_s if sub_rec else 0.0
    dist_score=1.0 if (dist_rec and (len(dist_recs)==1 or sub_rec)) else (0.6 if dist_rec else 0.0)
    conf=round(100*(0.34*prov_s+0.33*min(dist_score,1)+0.33*min(sub_score,1)))
    conflict_yes="Yes" if any(f.startswith("conflict:") for f in flags) else ""

    addr_rows.append(dict(house=house_col,moo=moo_col,soi=soi or "",road=road or "",street=street_part,
        sub=sub_th,sub_en=sub_en,dist=dist_th,dist_en=dist_en,prov=prov_native,prov_en=prov_en,
        postal=postal or "",tid=tid,did=did,pid=pid,full_th=full_th,full_en=full_en,
        status=status,conf=conf,source=source,conflict=conflict_yes,
        flags="; ".join(flags) if flags else "ok"))

addr = pd.DataFrame(addr_rows, index=sap.index)
addmap={"sap_addr_house":"house","sap_addr_moo":"moo","sap_addr_soi":"soi","sap_addr_road":"road",
    "sap_addr_street":"street","sap_addr_subdistrict":"sub","sap_addr_district":"dist","sap_addr_province":"prov",
    "sap_addr_subdistrict_en":"sub_en","sap_addr_district_en":"dist_en","sap_addr_province_en":"prov_en",
    "sap_addr_postal":"postal","sap_tambon_id":"tid","sap_district_id":"did","sap_province_id":"pid",
    "sap_addr_full_th":"full_th","sap_addr_full_en":"full_en","sap_addr_status":"status",
    "sap_addr_confidence":"conf","sap_addr_source":"source","address_conflict":"conflict","sap_addr_flags":"flags"}
for col,src in addmap.items(): sap[col]=addr[src].values
st=Counter(addr['status'])
print(f"  address status: {dict(st)}")

# ══════════════════════════════════════════════════════════════════════════════
# PART D — build indices for matching (master name/address, maskey org/cuscode)
# ══════════════════════════════════════════════════════════════════════════════
print("Building match indices …")
mas['_nn']=mas['NAME1'].apply(norm_name)
name_counts=Counter(mas['_nn'])                      # >1 = ชื่อนี้ซ้ำหลาย Cuscode ใน master เอง (generic)
name_to_cuscodes=defaultdict(list)
for _,r in mas.iterrows():
    if r['_nn']: name_to_cuscodes[r['_nn']].append(r['_k'])
NAME_BLOCK=defaultdict(list)                         # blocking: first-4-char prefix -> [(norm_name,cuscode)]
for nn,k in zip(mas['_nn'],mas['_k']):
    if nn: NAME_BLOCK[nn[:4]].append((nn,k))

maskey_rows=sap[sap['maskey_status']=='MASKEY']
cuscode_to_maskey=defaultdict(list)                  # cuscode -> [(LocationID, sap_code_str), ...]
for _,r in maskey_rows.iterrows():
    cuscode_to_maskey[r['_k']].append((r['Location ID'], r['SVMX_SAP_Code__c']))
cuscode_has_maskey=set(cuscode_to_maskey.keys())

org_to_cuscodes=defaultdict(set)                     # Organization ID -> {cuscodes จาก maskey ในนั้น}
for _,r in maskey_rows.iterrows():
    if pd.notna(r['Organization ID']): org_to_cuscodes[r['Organization ID']].add(r['_k'])
n_org_conflict=sum(1 for v in org_to_cuscodes.values() if len(v)>1)
print(f"  master names: {len(name_to_cuscodes):,} distinct | generic (>1 cuscode): {sum(1 for c in name_counts.values() if c>1):,}")
print(f"  org groups with maskey: {len(org_to_cuscodes):,} | conflicting (>1 distinct cuscode): {n_org_conflict}")

def addr_score_for(sap_row, cuscode):
    """M4: เทียบ ID ทางการ (ไม่ fuzzy text) sap_tambon/district/province_id vs master addr_*_id."""
    mr=mas_by_k.get(cuscode)
    if mr is None: return None
    tid,did,pid = sap_row['sap_tambon_id'], sap_row['sap_district_id'], sap_row['sap_province_id']
    mtid,mdid,mpid = mr.get('addr_tambon_id'), mr.get('addr_district_id'), mr.get('addr_province_id')
    if not any([tid,did,pid]) or not any([mtid,mdid,mpid]): return None   # ไม่มีข้อมูลพอเทียบ = unknown ไม่ใช่ไม่ตรง
    if tid and mtid and tid==mtid: return 1.0
    if did and mdid and did==mdid: return 0.6
    if pid and mpid and pid==mpid: return 0.3
    return 0.0

# ══════════════════════════════════════════════════════════════════════════════
# PART E — per-row decision: M1(name) + M2(org) + M4(address-ID) -> action
# ══════════════════════════════════════════════════════════════════════════════
def best_fuzzy_name(nn):
    """M1 fallback: fuzzy ภายใน block (first-4-char) เท่านั้น กันช้าเกินไป."""
    best=None; bscore=0.0
    for cand_nn,cus in NAME_BLOCK.get(nn[:4],[]):
        s=difflib.SequenceMatcher(None,nn,cand_nn).ratio()
        if s>bscore: bscore,best=s,cus
    return (best,bscore) if best and bscore>=NAME_FLOOR else (None,0.0)

def score_of(c):
    s=0.0
    if c['org']: s+=50
    if c['name_score']: s+=c['name_score']*30
    if c['addr_score']: s+=c['addr_score']*20
    return s

def decide(row):
    org_id=row.get('Organization ID')
    candidates={}   # cuscode -> dict(org,name_score,name_method,generic)
    def touch(cus):
        return candidates.setdefault(cus, dict(org=False,name_score=None,name_method=None,generic=False))

    org_targets=org_to_cuscodes.get(org_id, set()) if pd.notna(org_id) else set()
    org_conflict=len(org_targets)>1
    for cus in org_targets:              # ทุก Cuscode ที่มี maskey ใน org นี้ = candidate ที่ valid ในตัวเอง
        c=touch(cus); c['org']=True      # (sap code ตรง master คือหลักฐานหนักแล้ว ไม่ทิ้งแม้ org มีหลาย Cuscode)

    for nm,src in [(row.get('Location Name'),'loc'),(row.get('Organization name'),'org')]:
        nn=norm_name(nm)
        if not nn: continue
        exact=name_to_cuscodes.get(nn)
        if exact:
            generic=len(exact)>1
            for cus in exact:
                if src=='org' and cus in org_targets:
                    continue   # Organization name ตรงกับ maskey ของ org ตัวเอง = ข้อมูลวนซ้ำ (org field ถูกกำหนดเหมือนกันทั้ง org อยู่แล้ว) ไม่ใช่หลักฐานอิสระใหม่ ข้าม
                c=touch(cus)
                if c['name_score'] is None or 1.0>c['name_score']:
                    c['name_score']=1.0; c['name_method']=f'name_exact({src})'
                c['generic']=c['generic'] or generic
        else:
            cus,sc=best_fuzzy_name(nn)
            if cus and not (src=='org' and cus in org_targets):
                c=touch(cus)
                if c['name_score'] is None or sc>c['name_score']:
                    c['name_score']=sc; c['name_method']=f'name_fuzzy({src},{sc:.2f})'
                c['generic']=c['generic'] or (name_counts.get(nn,0)>1)

    if not candidates:
        return dict(action='NEW', target=None, method='', score=0, review='')

    for cus,c in candidates.items():
        c['addr_score']=addr_score_for(row,cus)

    ranked=sorted(candidates.items(), key=lambda kv:-score_of(kv[1]))
    top_cus,top=ranked[0]; top_score=score_of(top)

    tier='REVIEW'
    if top['org'] and not org_conflict:
        tier='AUTO'
    elif (top['name_score'] or 0)>=NAME_STRONG and not top['generic']:
        tier='AUTO'
    elif (top['name_score'] or 0)>=NAME_FLOOR and (top['addr_score'] or 0)>=0.6:
        tier='AUTO'   # ชื่อกลาง/generic ต้องมีที่อยู่ยืนยันระดับอำเภอขึ้นไปเสมอ

    if len(ranked)>1 and tier=='AUTO' and not (top['org'] and not org_conflict):
        gap=top_score-score_of(ranked[1][1])
        if gap<15 and ranked[1][0]!=top_cus:
            tier='REVIEW'   # candidate สูสีกันเกินไป ไม่ชัวร์พอ auto (ยกเว้น org_sibling ที่ไม่ conflict — เชื่อได้ ไม่ต้อง tie-break)

    methods=[]
    if top['org']: methods.append('org_sibling'+('[CONFLICT]' if org_conflict else ''))
    if top['name_method']: methods.append(top['name_method'])
    if top['addr_score']: methods.append(f"addr({top['addr_score']:.1f})")

    if tier=='AUTO':
        action='MERGE_WITH' if top_cus in cuscode_has_maskey else 'ASSIGN_CODE'
        review=''
    elif tier=='REVIEW':
        action='REVIEW'
        review='; '.join(f"{cus}:{mas_by_k[cus]['NAME1']}(score={score_of(c):.0f})" for cus,c in ranked[:3])
        top_cus=None   # REVIEW ต้องดูจาก review_candidates (มีหลายตัวเลือก) ไม่ใช่ target เดียวที่ทำให้ดูมั่นใจเกินจริง
    else:
        action='NEW'; review=''
        top_cus=None

    return dict(action=action, target=top_cus, method='+'.join(methods), score=round(top_score), review=review)

print("Matching non-maskey rows …")
nonmaskey_idx = sap.index[sap['maskey_status']!='MASKEY']
match_rows={}
for n,i in enumerate(nonmaskey_idx):
    match_rows[i]=decide(sap.loc[i])
    if (n+1)%2000==0: print(f"  {n+1:,}/{len(nonmaskey_idx):,} …")
print(f"  done: {len(match_rows):,} non-maskey rows matched")
print("  action distribution:", Counter(v['action'] for v in match_rows.values()))

# ══════════════════════════════════════════════════════════════════════════════
# PART E.2 — DELETE rule: REVIEW/NEW (หา maskey ให้ MERGE_WITH ไม่ได้) + ไม่มีทั้ง WO และ IB
# ไม่แตะ ASSIGN_CODE (รู้ตัวตนจาก master แล้ว แค่ไม่มี maskey ให้ join — เก็บไว้)
# ══════════════════════════════════════════════════════════════════════════════
print("Applying DELETE rule (REVIEW/NEW + ไม่มีทั้ง WO และ IB) …")
no_wo_ib = sap['No. of Work order'].isna() & sap['No. of IB'].isna()
n_deleted=0
for i,v in match_rows.items():
    if v['action'] in ('REVIEW','NEW') and no_wo_ib[i]:
        v['method'] = (v['method']+'+' if v['method'] else '') + 'no_WO_no_IB'
        v['action'] = 'DELETE'
        n_deleted+=1
print(f"  DELETE: {n_deleted:,} แถว (จาก REVIEW/NEW ที่ไม่มีทั้ง WO และ IB)")
print("  action distribution (หลัง DELETE rule):", Counter(v['action'] for v in match_rows.values()))

# ══════════════════════════════════════════════════════════════════════════════
# PART F — maskey-side info (Step-1 equivalent, own SAP code) + assemble output
# ══════════════════════════════════════════════════════════════════════════════
print("Computing maskey-side master info …")
code_counts = sap.loc[sap['maskey_status']=='MASKEY','_k'].value_counts()
mk_info={}
for i in sap.index[sap['maskey_status']=='MASKEY']:
    row=sap.loc[i]; mr=mas_by_k[row['_k']]
    s1=name_sim(mr['NAME1'], row.get('Location Name')); s2=name_sim(mr['NAME1'], row.get('Organization name'))
    cand=[x for x in (s1,s2) if x is not None]; best=max(cand) if cand else None
    nchk="OK" if (best is None or best>=NAME_FLOOR) else f"NAME_MISMATCH (sim={best:.2f})"
    cnt=int(code_counts.get(row['_k'],1))
    multi=f"SHARED_SAP_CODE: code มีใน master แต่ผูกกับ {cnt} locations" if cnt>1 else ""
    mk_info[i]=dict(cuscode=mr['Cuscode'], name=mr['NAME1'], fth=mr['addr_full_th'], fen=mr['addr_full_en'],
                     mstat=mr['addr_status'], nchk=nchk, multi=multi)

print("Assembling final combined output …")
out = sap.drop(columns=['_k']).copy()
out['sap_match_status'] = sap['maskey_status'].replace({'MASKEY':'MASKEY (matched to master)'})

# maskey-side (Step-1 equivalent) columns
out['master_cuscode']=""; out['master_name']=""; out['master_addr_full_th']=""; out['master_addr_full_en']=""
out['master_addr_status']=""; out['name_check']=""; out['multi_location_note']=""
for i,d in mk_info.items():
    out.at[i,'master_cuscode']=d['cuscode']; out.at[i,'master_name']=d['name']
    out.at[i,'master_addr_full_th']=d['fth']; out.at[i,'master_addr_full_en']=d['fen']
    out.at[i,'master_addr_status']=d['mstat']; out.at[i,'name_check']=d['nchk']; out.at[i,'multi_location_note']=d['multi']

# Stage-2 matching columns (non-maskey rows only; blank for maskey)
for col in ['match_action','match_target_cuscode','match_target_name','match_target_location_ids',
            'match_target_addr_full_th','match_target_addr_full_en','match_method','match_score','review_candidates']:
    out[col]=""
out.loc[sap['maskey_status']=='MASKEY','match_action']="MASKEY"
for i,v in match_rows.items():
    out.at[i,'match_action']=v['action']
    out.at[i,'match_method']=v['method']
    out.at[i,'match_score']=str(v['score'])
    out.at[i,'review_candidates']=v['review']
    if v['target'] is not None:
        mr=mas_by_k[v['target']]
        out.at[i,'match_target_cuscode']=mr['Cuscode'].lstrip('0') or '0'   # SAP format ตัด 0 นำหน้า
        out.at[i,'match_target_name']=mr['NAME1']
        out.at[i,'match_target_addr_full_th']=mr['addr_full_th']
        out.at[i,'match_target_addr_full_en']=mr['addr_full_en']
        if v['action']=='MERGE_WITH':
            locs=cuscode_to_maskey.get(v['target'],[])
            out.at[i,'match_target_location_ids']=", ".join(f"{lid}({sc})" for lid,sc in locs)

print("Writing workbook …")
ADDED_ADDR=list(addmap.keys())
ADDED_MASKEY=['sap_match_status','master_cuscode','master_name','master_addr_full_th','master_addr_full_en',
              'master_addr_status','name_check','multi_location_note']
ADDED_S2=['match_action','match_target_cuscode','match_target_name','match_target_location_ids',
          'match_target_addr_full_th','match_target_addr_full_en','match_method','match_score','review_candidates']

FILL_ADDR={"VERIFIED":PatternFill("solid",start_color="C8E6C9"),"AUTO_FIXED":PatternFill("solid",start_color="BBDEFB"),
      "FUZZY_FIXED":PatternFill("solid",start_color="FFF9C4"),"NEEDS_REVIEW":PatternFill("solid",start_color="FFCDD2")}
FILL_ACTION={"MASKEY":PatternFill("solid",start_color="B0BEC5"),"MERGE_WITH":PatternFill("solid",start_color="C8E6C9"),
      "ASSIGN_CODE":PatternFill("solid",start_color="A5D6A7"),"REVIEW":PatternFill("solid",start_color="FFF9C4"),
      "NEW":PatternFill("solid",start_color="E0E0E0"),"DELETE":PatternFill("solid",start_color="EF9A9A")}
RED=PatternFill("solid",start_color="FFCDD2"); ORANGE=PatternFill("solid",start_color="FFE0B2")

wb=Workbook(); del wb["Sheet"]; ws=wb.create_sheet("Locations_Matched")
cols=list(out.columns)
addr_set={cols.index(c)+1 for c in ADDED_ADDR}; mk_set={cols.index(c)+1 for c in ADDED_MASKEY}; s2_set={cols.index(c)+1 for c in ADDED_S2}
astat_col=cols.index("sap_addr_status")+1; act_col=cols.index("match_action")+1
nchk_col=cols.index("name_check")+1; multi_col=cols.index("multi_location_note")+1
data=[[v if pd.notna(v) else None for v in r] for r in out.itertuples(index=False)]
allrows=[cols]+data
for ri,row in enumerate(allrows,1):
    for ci,val in enumerate(row,1):
        c=ws.cell(ri,ci,val)
        if ri==1:
            c.font=Font(bold=True,color="FFFFFF",name="Arial",size=10)
            fill=PatternFill("solid",start_color="F57C00") if ci in addr_set else \
                 PatternFill("solid",start_color="6A1B9A") if ci in mk_set else \
                 PatternFill("solid",start_color="00695C") if ci in s2_set else \
                 PatternFill("solid",start_color="37474F")
            c.fill=fill; c.alignment=Alignment(horizontal="center",wrap_text=True)
        else:
            c.font=Font(name="Arial",size=10)
            if ci==astat_col and val in FILL_ADDR: c.fill=FILL_ADDR[val]
            elif ci==act_col and val in FILL_ACTION: c.fill=FILL_ACTION[val]
            elif ci==nchk_col and isinstance(val,str) and val.startswith("NAME_MISMATCH"): c.fill=RED
            elif ci==multi_col and val: c.fill=ORANGE
ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
for ci in range(1,len(cols)+1):
    w=max((len(str(r[ci-1] or "")) for r in allrows[:60]),default=10)
    ws.column_dimensions[ws.cell(1,ci).column_letter].width=min(w+2,46)

# Summary
N=len(out); action_ct=Counter(out['match_action'])
ws2=wb.create_sheet("Summary")
rowsS=[["SAP <-> Master — Stage 2 Multi-Signal Matching — Summary",""],["",""],
       ["match_action (ทุก 18,007 แถว)","จำนวน","%"]]
for k in ["MASKEY","MERGE_WITH","ASSIGN_CODE","REVIEW","NEW","DELETE"]:
    rowsS.append([k, action_ct.get(k,0), f"{100*action_ct.get(k,0)/N:.1f}%"])
rowsS+=[["รวม",N,"100%"],["",""],
    ["เฉพาะ non-maskey (8,836 แถว ที่พยายาม match)","",""],
    ["  MERGE_WITH + ASSIGN_CODE (auto)", action_ct.get("MERGE_WITH",0)+action_ct.get("ASSIGN_CODE",0),
     f"{100*(action_ct.get('MERGE_WITH',0)+action_ct.get('ASSIGN_CODE',0))/8836:.1f}%"],
    ["  REVIEW", action_ct.get("REVIEW",0), f"{100*action_ct.get('REVIEW',0)/8836:.1f}%"],
    ["  NEW", action_ct.get("NEW",0), f"{100*action_ct.get('NEW',0)/8836:.1f}%"],
    ["",""],
    ["หมายเหตุ REVIEW แบ่งตามสาเหตุ",""],
    ["  org_conflict (org เดียวกันมี maskey ชี้คนละ Cuscode)", n_org_conflict, "master อาจมี Cuscode ซ้ำสำหรับบริษัทเดียวกัน"],
    ["  weak/tied signal", action_ct.get("REVIEW",0)-n_org_conflict, "ชื่อ/ที่อยู่ไม่ชัดพอ หรือ candidate สูสีกัน"]]
for ri,row in enumerate(rowsS,1):
    for ci,val in enumerate(row,1):
        c=ws2.cell(ri,ci,val)
        if ri==1: c.font=Font(bold=True,size=13)
        if ci==1 and val in FILL_ACTION: c.fill=FILL_ACTION[val]
ws2.column_dimensions['A'].width=46; ws2.column_dimensions['B'].width=12; ws2.column_dimensions['C'].width=50

# Notes
ws3=wb.create_sheet("Notes")
def put(r,t,b=False):
    c=ws3.cell(r,1,t); c.font=Font(bold=b,name="Arial",size=12 if (b and r==1) else (11 if b else 10)); c.alignment=Alignment(wrap_text=True,vertical="top")
r=1
for t,b in [("SAP ↔ Master — Stage 2: Multi-Signal Matching — วิธีการ",True),("",False),
 ("★ SELF-CONTAINED: คำนวณ maskey-join (Step1) + address-clean (Stage1) ซ้ำจาก 3 ไฟล์ต้นทางในสคริปต์นี้เอง",True),
 ("  ไม่อ่าน output ไฟล์ก่อนหน้า กันไฟล์เก่าเพี้ยนถ้า logic เคยแก้ไปแล้วแต่ยังไม่ re-run",False),
 ("",False),("3 สัญญาณที่ใช้ (ไม่ใช้ dedup key เดิมของ ServiceMax — Location Duplication Check/Clean street/Suspect ตามที่ตกลง)",True),
 ("M1 ชื่อ — Location/Organization name (normalize ตัดคำต่อท้ายบริษัท) เทียบ master NAME1: exact-match ก่อน, fuzzy ใน block (first-4-char) ถ้าไม่ตรง",False),
 ("M2 Organization — non-maskey อยู่ Organization ID เดียวกับ maskey ไหม เสนอ 'ทุก' Cuscode ที่มี maskey ใน org นั้นเป็น candidate (แม้ org จะมีมากกว่า 1 Cuscode) ให้ชื่อ/ที่อยู่ช่วยเลือกตัวที่ใช่",False),
 ("",False),("★ พบระหว่างตรวจ (2 รอบ): Organization ID บางอัน (114 org) เป็น 'ถังขยะ' รวมบริษัทไม่เกี่ยวกันหลายสิบเจ้าไว้ผิด ๆ",True),
 ("  เช่น org หนึ่งชื่อ 'MIT Technology Co., Ltd.' มี 30 location ที่จริงเป็นคนละบริษัท (มหิดล, KCE Electronics, THAI PREEDA ฯลฯ)",False),
 ("  แก้: 'Organization name' ที่ตรงกับ Cuscode ของ maskey ใน org ตัวเอง = ข้อมูลวนซ้ำ (field นี้ถูกกำหนดเหมือนกันทั้ง org) ไม่ใช่หลักฐานอิสระ ตัดออกจาก M1",False),
 ("  แถวกลุ่มนี้ตอนนี้ลง REVIEW ให้คนตรวจ (ก่อนแก้ เคยหลุดไป MERGE_WITH ผิดบริษัทเพราะชื่อ org วนกลับมาชนกับ org signal เอง)",False),
 ("M4 ที่อยู่ — เทียบ ID ทางการ (sap_tambon/district/province_id vs master addr_*_id) ไม่ fuzzy text; ไม่มีข้อมูลให้เทียบ = 'ไม่ทราบ' ไม่ใช่ 'ไม่ตรง'",False),
 ("",False),("เกณฑ์ auto (MERGE_WITH/ASSIGN_CODE) — ป้องกัน false merge",True),
 ("• org_sibling ไม่ conflict → auto ทันที (เชื่อได้สุด ไม่ต้องมีสัญญาณอื่นเสริม)",False),
 ("• ชื่อตรง unique (ไม่ generic คือไม่ซ้ำ Cuscode อื่นใน master) → auto",False),
 ("• ชื่อกลาง/generic → ต้องมีที่อยู่ยืนยันระดับอำเภอขึ้นไปเสมอ (addr_score>=0.6) ถึง auto",False),
 ("• candidate สูสีกันเกินไป (ต่างกัน<15 คะแนน) → REVIEW เสมอ ยกเว้น org_sibling ไม่ conflict",False),
 ("",False),("MERGE_WITH vs ASSIGN_CODE",True),
 ("Cuscode ปลายทางมี maskey (location ที่มี SAP code ตรงอยู่แล้ว) → MERGE_WITH = Location ID(SAP code)",False),
 ("Cuscode ปลายทางไม่มี maskey เลย → ASSIGN_CODE = Cuscode ตัด 0 นำหน้า (SAP format)",False),
 ("",False),("★ พบระหว่างตรวจ: 1,435 org (24%) มี maskey ชี้คนละ Cuscode ที่ 'ถูกทั้งคู่' เช่น 3M/24K มี 2 Cuscode ใน master",True),
 ("  → เป็นไปได้ว่า master มี Cuscode ซ้ำสำหรับบริษัทเดียวกัน (ปัญหาลูกค้าซ้ำ ไม่ใช่แค่ location ซ้ำ) — ไม่แก้ในรอบนี้ REVIEW ไว้ให้ตรวจ",False),
 ("",False),("คอลัมน์ที่เพิ่ม (Stage 2)",True),
 ("match_action — MASKEY/MERGE_WITH/ASSIGN_CODE/REVIEW/NEW (สี)",False),
 ("match_target_cuscode/name/addr_full_th/en — เป้าหมายที่ match ได้ (ที่อยู่ clean พร้อมใช้)",False),
 ("match_target_location_ids — location(s) ที่มี maskey อยู่แล้วที่ Cuscode นี้ (เฉพาะ MERGE_WITH)",False),
 ("match_method — สัญญาณที่ใช้ตัดสิน (org_sibling/name_exact/name_fuzzy/addr)",False),
 ("match_score / review_candidates — คะแนน / รายชื่อ candidate ให้คนตรวจ (เฉพาะ REVIEW)",False),
 ("",False),("★ DELETE rule: REVIEW/NEW (หา maskey ให้ MERGE_WITH ไม่ได้) + ไม่มีทั้ง No. of Work order และ No. of IB เลย -> DELETE",True),
 ("  ไม่แตะ ASSIGN_CODE (รู้ตัวตนจาก master แล้วแม้ไม่มี maskey — เก็บไว้เสมอ) และไม่แตะ MASKEY/MERGE_WITH อยู่แล้ว",False),
 ("  เหตุผล: ไม่มีทั้งประวัติงานซ่อม (WO) และอุปกรณ์ติดตั้ง (IB) = ไม่เคยมีกิจกรรมจริง ต่อให้หาเจ้าของไม่ได้ก็ไม่เสียข้อมูลสำคัญถ้าลบ",False),
 ("  หมายเหตุ: ค่า WO/IB ที่ 'ไม่มี' เก็บเป็นค่าว่าง (blank) ในไฟล์ต้นฉบับ ไม่ใช่เลข 0 — เช็คด้วยเงื่อนไข 'ว่างทั้งคู่'",False),
 ("",False),("ขอบเขต: จับคู่ non-maskey ↔ master เท่านั้น (ยังไม่ทำ duplicate-clustering ระหว่าง non-maskey กันเอง — เลื่อนไป stage ถัดไป)",True),
 ("สถานะสี match_action: เทา MASKEY · เขียว MERGE_WITH/ASSIGN_CODE · เหลือง REVIEW · เทาอ่อน NEW",True)]:
    put(r,t,b); r+=1
ws3.column_dimensions['A'].width=100
wb.save(OUT)

print("\n"+"="*58+"\n  STAGE 2 COMPLETE")
print(f"  Total: {N:,} | non-maskey attempted: {len(match_rows):,}")
for k in ["MASKEY","MERGE_WITH","ASSIGN_CODE","REVIEW","NEW","DELETE"]:
    print(f"    {k:12} {action_ct.get(k,0):6,} ({100*action_ct.get(k,0)/N:4.1f}%)")
print(f"  Output: {OUT.name}\n"+"="*58)