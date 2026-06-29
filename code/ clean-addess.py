"""
TH44 Address Cleaning V2 — Authoritative-Gazetteer Engine
=========================================================
Database-driven (Thai Postal V2), NO guessing, NO LLM.
- Province  : authoritative lookup from POSTAL CODE (98.6% 1:1)
- District  : implied by matched tambon, or postal (single), else matched
- Subdistrict: in-house n-gram + cosine matcher over postal-narrowed candidates
- Wrong-column: matcher scans a field pool; new value written to the CORRECT column
                (raw fields are never modified)
- Output    : bilingual (TH official + EN official), colour-coded status,
              confidence = real match score, every change flagged.
Pure stdlib + pandas + openpyxl (no network, no external model).
"""
import re, math
from pathlib import Path
from collections import Counter, defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

DB    = Path("/Users/phachon/Documents/DKSH/location-clean-up/input/thai-postal-codes_V2.xlsx")
DATA  = Path("/Users/phachon/Documents/DKSH/location-clean-up/input/TH44_all_CUSTOMERS_excl_flag_del.xlsx")
OUT   = Path("/Users/phachon/Documents/DKSH/location-clean-up/output/TH44_Addresses_Cleaned_V2.xlsx")

ACCEPT, EXACT = 0.60, 0.999          # match thresholds (tunable)

# ── helpers ────────────────────────────────────────────────────────────────────
def clean_ws(s):
    if pd.isna(s): return None
    s = re.sub(r"\s{2,}", " ", str(s).strip()).strip(" ,")
    return s or None

TH_RE = re.compile(r"[\u0E00-\u0E7F]")
def script_of(s):
    if not s: return None
    t = TH_RE.search(s); l = re.search(r"[A-Za-z]", s)
    if t and not l: return "TH"
    if l and not t: return "EN"
    if t and l: return "TH" if len(TH_RE.findall(s)) >= len(re.findall(r"[A-Za-z]", s)) else "EN"
    return None

PFX = re.compile(r"^(ตำบล|แขวง|ต\.|อำเภอ|เขต|อ\.|จังหวัด|จ\.|tambon|khwaeng|amphoe|khet|"
                 r"sub-?district|district|province)\s*", re.I)
def norm_match(s):
    s = clean_ws(s)
    if not s: return None
    prev = None
    while prev != s:
        prev = s; s = PFX.sub("", s).strip()
    if re.search(r"[A-Za-z]", s):
        return re.sub(r"[^a-z0-9]", "", s.lower())   # romanisation-robust (drop spaces/punct)
    return s

def parse_house(s):
    s = clean_ws(s)
    if not s: return None, None
    s2 = re.sub(r"^เลขที่\s*", "", s).strip()
    if re.search(r"\d", s2):
        m = re.match(r"^([\d/\-\s]+)", s2)
        return (m.group(1).strip() if m else s2), None
    return None, s2
def parse_moo(s):
    s = clean_ws(s)
    if not s: return None
    m = re.search(r"(?:หมู่ที่|หมู่|ม\.)\s*(\d+)", s)
    return m.group(1) if m else (s if re.fullmatch(r"\d+", s) else None)
def norm_road(s):
    s = clean_ws(s)
    if not s: return None
    s = re.sub(r"^ถ\.\s*", "ถนน", s); s = re.sub(r"^ซ\.\s*", "ซอย", s)
    return s

# ── in-house n-gram cosine matcher ─────────────────────────────────────────────
class NGramCosine:
    def __init__(self, ns=(2, 3)): self.ns = ns; self.idf = {}; self.dflt = 1.0; self.vecs = {}
    def _grams(self, s):
        s = f"  {s}  "; out = []
        for n in self.ns: out += [s[i:i+n] for i in range(len(s) - n + 1)]
        return out
    def fit(self, names):
        names = {n for n in names if n}
        dfc = Counter()
        for nm in names:
            for g in set(self._grams(nm)): dfc[g] += 1
        N = len(names)
        self.idf = {g: math.log((N + 1) / (d + 1)) + 1 for g, d in dfc.items()}
        self.dflt = math.log(N + 1) + 1
        for nm in names: self.vecs[nm] = self._vec(nm)
    def _vec(self, s):
        c = Counter(self._grams(s))
        v = {g: cnt * self.idf.get(g, self.dflt) for g, cnt in c.items()}
        nrm = math.sqrt(sum(w * w for w in v.values())) or 1.0
        return {g: w / nrm for g, w in v.items()}
    def _sim(self, qv, name):
        cv = self.vecs.get(name) or self._vec(name)
        a, b = (qv, cv) if len(qv) < len(cv) else (cv, qv)
        return sum(w * b.get(g, 0.0) for g, w in a.items())
    def best(self, query, candidates):
        q = norm_match(query)
        if not q or not candidates: return None, 0.0
        qv = self._vec(q)
        s = max(((self._sim(qv, c), c) for c in candidates), default=(0.0, None))
        return s[1], round(s[0], 3)

# ── load DB + build authoritative indices ──────────────────────────────────────
print("Loading V2 database …")
xl = pd.ExcelFile(DB)
post = xl.parse("รหัสไปรษณีย์", dtype=str)
tam  = xl.parse("TambonDatabase", dtype=str)
tam_ix = tam.set_index("TambonID")

def rec(tid):
    r = tam_ix.loc[tid]
    pid = r.ProvinceID
    bkk = pid == "10"
    return dict(tid=tid, did=r.DistrictID, pid=pid,
                sub_th=("แขวง" if bkk else "ตำบล") + r.TambonThaiShort, sub_sh=r.TambonThaiShort,
                sub_en=r.TambonEngShort, dist_th=r.DistrictThai, dist_sh=r.DistrictThaiShort,
                dist_en=r.DistrictEng, dist_ensh=r.DistrictEngShort,
                prov_th=r.ProvinceThai, prov_en=r.ProvinceEng, bkk=bkk)

postal_tambons = defaultdict(list); postal_prov = defaultdict(Counter)
for _, row in post.iterrows():
    pc, tid = clean_ws(row.PostCode), row.TambonID
    if pc and tid in tam_ix.index:
        rc = rec(tid); postal_tambons[pc].append(rc); postal_prov[pc][rc["prov_th"]] += 1

# province resolver (City cross-check + fallback)
PROV_TH = set(tam["ProvinceThai"].dropna())
PROV_EN_NORM = {e.lower(): t for e, t in zip(tam["ProvinceEng"], tam["ProvinceThai"]) if pd.notna(e)}
PROV_LIST = list(PROV_TH)
BKK_ALIAS = {"กรุงเทพ","กรุงเทพฯ","กทม","กทม.","bangkok","krungthep","bkk"}
PROV_ALIAS = {"ayutthaya":"พระนครศรีอยุธยา","korat":"นครราชสีมา","sriracha":"ชลบุรี"}
import difflib
def resolve_prov(s):
    s = clean_ws(s)
    if not s: return None
    b = re.sub(r"\s*\bprovince\b\s*", "", s, flags=re.I)
    b = re.sub(r"^(จังหวัด|จ\.|จ\s)\s*", "", b).strip()
    low = re.sub(r"\s+", "", b.lower()).replace("ฯ", "")
    if low in BKK_ALIAS or b in {"กรุงเทพ","กรุงเทพฯ"}: return "กรุงเทพมหานคร"
    if low in PROV_ALIAS: return PROV_ALIAS[low]
    if b in PROV_TH: return b
    if low in PROV_EN_NORM: return PROV_EN_NORM[low]
    if TH_RE.search(b):
        m = difflib.get_close_matches(b, PROV_LIST, n=1, cutoff=0.82)
        if m: return m[0]
    else:
        m = difflib.get_close_matches(low, list(PROV_EN_NORM), n=1, cutoff=0.82)
        if m: return PROV_EN_NORM[m[0]]
    return None

# fallback hierarchy indices
prov_tambons = defaultdict(list); prov_districts = defaultdict(dict)
for tid in tam_ix.index:
    rc = rec(tid); prov_tambons[rc["prov_th"]].append(rc)
    prov_districts[rc["prov_th"]].setdefault(rc["did"], rc)

# matcher fit on all official short names (TH + EN), normalised consistently
_allnames = set()
for _col in ["TambonThaiShort", "TambonEngShort", "DistrictThaiShort", "DistrictEngShort"]:
    for _v in tam[_col].dropna():
        _nm = norm_match(_v)
        if _nm: _allnames.add(_nm)
M = NGramCosine(); M.fit(_allnames)
print(f"  {len(postal_tambons)} postal codes | matcher vocab {len(M.idf):,}")

# ── matching helpers ───────────────────────────────────────────────────────────
def best_rec(query, recs, level, sc):
    """level in {'sub','dist'}; returns (rec, score)."""
    if not query or not recs: return None, 0.0
    key = ({"sub":"sub_sh","dist":"dist_sh"} if sc == "TH" else {"sub":"sub_en","dist":"dist_ensh"})[level]
    name2rec = {}
    for r in recs:
        nm = norm_match(r[key])
        if nm: name2rec.setdefault(nm, r)
    name, score = M.best(query, list(name2rec))
    return (name2rec.get(name), score) if name else (None, 0.0)

def best_pool(fields, recs, level, sc):
    """fields = list of (fieldname, value); pick highest-scoring source. Returns (rec, score, src).
    Script is detected per value so mixed-script rows match the correct candidate names."""
    br, bs, bsrc = None, 0.0, None
    for fname, val in fields:
        sv = script_of(val) or sc
        r, s = best_rec(val, recs, level, sv)
        if s > bs: br, bs, bsrc = r, s, fname
    return br, bs, bsrc

# ── status colours ─────────────────────────────────────────────────────────────
STATUS_FILL = {
    "VERIFIED":     PatternFill("solid", start_color="C8E6C9"),
    "AUTO_FIXED":   PatternFill("solid", start_color="BBDEFB"),
    "FUZZY_FIXED":  PatternFill("solid", start_color="FFF9C4"),
    "NEEDS_REVIEW": PatternFill("solid", start_color="FFCDD2"),
    "FOREIGN":      PatternFill("solid", start_color="E0E0E0"),
}
RANK = {"VERIFIED":0, "AUTO_FIXED":1, "FUZZY_FIXED":2, "NEEDS_REVIEW":3, "FOREIGN":4}

# ── per-row resolution ─────────────────────────────────────────────────────────
print("Resolving …")
df = pd.read_excel(DATA, dtype=str)
N = len(df)
rows = []

def disp_prov_th(rc_or_provth):
    p = rc_or_provth
    return "กรุงเทพมหานคร" if p == "กรุงเทพมหานคร" else f"จังหวัด{p}"

for i in df.index:
    g = {c: clean_ws(df.at[i, c]) for c in
         ["STREET","STREET2","STREET3","STREET4","STREET5","Dist","Other city","City","POSTAL CODE"]}
    sc = script_of(" ".join(filter(None, [g["Other city"], g["City"], g["Dist"]]))) or "TH"
    flags, comps = [], []
    postal = g["POSTAL CODE"] if (g["POSTAL CODE"] and re.fullmatch(r"\d{5}", g["POSTAL CODE"])) else None

    # FOREIGN
    if (df.at[i, "COUNTRY"] or "") != "Thailand":
        parts = [g[c] for c in ["STREET","STREET2","STREET3","STREET4","STREET5","Dist","Other city","City"]]
        native = ", ".join([p for p in parts if p])
        rows.append(dict(house="",bld="",moo="",soi="",road="",sub="",sub_en="",dist="",dist_en="",
                         prov=g["City"] or "", prov_en="", postal=postal or "", country=df.at[i,"COUNTRY"],
                         tid="",did="",pid="", full_th=native, full_en=native, status="FOREIGN",
                         conf=60, lang=sc or "EN", flags="foreign address — assembled as-is"))
        continue

    covered = postal in postal_tambons
    cands = postal_tambons.get(postal, [])

    # ---- PROVINCE (authoritative) ----
    if covered:
        prov_th = postal_prov[postal].most_common(1)[0][0]
    else:
        prov_th = resolve_prov(g["City"])
        flags.append(f"postal_not_in_db({postal})" if postal else "no_postal")
    city_key = resolve_prov(g["City"])
    prov_native = disp_prov_th(prov_th) if prov_th else (g["City"] or "")
    prov_en = (next((r["prov_en"] for r in cands), None)
               or (next((r["prov_en"] for r in prov_tambons.get(prov_th, [])), prov_th) if prov_th else ""))
    if not prov_th:
        comps.append("REVIEW")
    elif covered and city_key and city_key in PROV_TH and city_key != prov_th:
        comps.append("REVIEW"); flags.append(f"province:conflict(city={city_key}->postal={prov_th})")
    elif clean_ws(g["City"]) == prov_native:
        comps.append("UNCHANGED")
    else:
        comps.append("AUTO"); flags.append(f"province:standardized({g['City']}->{prov_native})")

    # candidate pools (district list for this postal)
    if covered:
        dist_recs = list({r["did"]: r for r in cands}.values())
    elif prov_th:
        dist_recs = list(prov_districts.get(prov_th, {}).values())
        cands = prov_tambons.get(prov_th, [])
    else:
        dist_recs = []

    # ---- SUBDISTRICT (tambon) via matcher over field pool ----
    sub_rec, sub_s, sub_src = best_pool(
        [("Dist", g["Dist"]), ("Other city", g["Other city"]), ("STREET4", g["STREET4"])], cands, "sub", sc)
    sub_present = any(g[f] for f in ["Dist", "Other city", "STREET4"])
    if sub_rec and sub_s >= ACCEPT:
        if sub_src != "Dist": flags.append(f"subdistrict:moved_from({sub_src})")
        raw_sub = norm_match(g["Dist"])
        if sub_s >= EXACT and raw_sub == sub_rec["sub_sh"] and sub_src == "Dist":
            comps.append("UNCHANGED")
        elif sub_s >= 0.95:
            comps.append("AUTO"); flags.append(f"subdistrict:standardized(->{sub_rec['sub_th']})")
        else:
            comps.append("FUZZY"); flags.append(f"subdistrict:fuzzy({g['Dist']}->{sub_rec['sub_sh']} {sub_s:.2f})")
    else:
        sub_rec = None
        if sub_present:
            flags.append(f"subdistrict:unmatched({g['Dist'] or g['Other city'] or g['STREET4']})")
            # not escalated to REVIEW here — district+province may still be solid (decided below)

    # ---- DISTRICT: tambon-implied > single-postal > matched ----
    if sub_rec:
        dist_rec = next((r for r in dist_recs if r["did"] == sub_rec["did"]), sub_rec)
        if sub_src != "Dist" or sub_s < 0.95: comps.append("AUTO")
    elif len(dist_recs) == 1:
        dist_rec = dist_recs[0]; comps.append("AUTO"); flags.append("district:from_postal(single)")
    else:
        dr, ds, dsrc = best_pool([("Other city", g["Other city"]), ("Dist", g["Dist"])], dist_recs, "dist", sc)
        if dr and ds >= ACCEPT:
            dist_rec = dr
            comps.append("AUTO" if ds >= 0.95 else "FUZZY")
            if ds < 0.95: flags.append(f"district:fuzzy({g['Other city']}->{dr['dist_sh']} {ds:.2f})")
        else:
            dist_rec = None
            if g["Other city"]: flags.append(f"district:unmatched({g['Other city']})"); comps.append("REVIEW")

    # ---- STREET DETAIL (raw untouched; new values to correct columns) ----
    house, bld = parse_house(g["STREET2"])
    if bld: flags.append(f"wrong_column:STREET2_is_name(->building)")
    moo = parse_moo(g["STREET3"])
    road = norm_road(g["STREET5"])
    if road and road != g["STREET5"]: flags.append("road:abbrev_expanded")
    soi = g["STREET4"] if (g["STREET4"] and re.search(r"ซอย|ซ\.|soi", g["STREET4"], re.I)
                           and (not sub_rec or sub_src != "STREET4")) else None

    # values
    sub_th = sub_rec["sub_th"] if sub_rec else ""
    sub_en = sub_rec["sub_en"] if sub_rec else ""
    dist_th = dist_rec["dist_th"] if dist_rec else ""
    dist_en = dist_rec["dist_ensh"] if dist_rec else ""
    tid = sub_rec["tid"] if sub_rec else ""
    did = dist_rec["did"] if dist_rec else ""
    pid = (cands[0]["pid"] if cands else (sub_rec["pid"] if sub_rec else ""))

    # ---- ASSEMBLE (full_th = always proper Thai; full_en = always English) ----
    seg = []
    if house: seg.append(f"เลขที่ {house}")
    if bld: seg.append(bld)
    if moo: seg.append(f"หมู่ที่ {moo}")
    if soi: seg.append(soi)
    if road: seg.append(road)
    if sub_th: seg.append(sub_th)
    if dist_th: seg.append(dist_th)
    if prov_native: seg.append(prov_native)
    if postal: seg.append(postal)
    full_th = " ".join(seg)
    head_en = [x for x in [house or bld, soi, road] if x]
    tail_en = [x for x in [sub_en, dist_en, prov_en, postal] if x]
    full_en = ", ".join(([" ".join(head_en)] if head_en else []) + tail_en + ["Thailand"])

    # ---- STATUS + CONFIDENCE ----
    order = [c for c in comps if c]
    if not dist_rec and not sub_rec and "REVIEW" not in order:
        order.append("REVIEW"); flags.append("incomplete:no_district_or_subdistrict")
    if "REVIEW" in order: status = "NEEDS_REVIEW"
    elif "FUZZY" in order: status = "FUZZY_FIXED"
    elif "AUTO" in order: status = "AUTO_FIXED"
    else: status = "VERIFIED"
    if not covered or not postal: status = "NEEDS_REVIEW"
    prov_s = 1.0 if (covered and "conflict" not in " ".join(flags)) else (0.5 if prov_th else 0.0)
    sub_score = sub_s if sub_rec else (1.0 if not sub_present else 0.0)
    dist_score = 1.0 if (dist_rec and (len(dist_recs) == 1 or sub_rec)) else (sub_s if dist_rec else 0.0)
    conf = round(100 * (0.34 * prov_s + 0.33 * min(dist_score, 1.0) + 0.33 * min(sub_score, 1.0)))

    rows.append(dict(house=house or "", bld=bld or "", moo=moo or "", soi=soi or "", road=road or "",
                     sub=sub_th, sub_en=sub_en, dist=dist_th, dist_en=dist_en,
                     prov=prov_native, prov_en=prov_en or "", postal=postal or "", country="Thailand",
                     tid=tid, did=did, pid=pid, full_th=full_th, full_en=full_en,
                     status=status, conf=conf, lang=sc, flags="; ".join(flags) if flags else "no_changes_needed"))

res = pd.DataFrame(rows, index=df.index)

# ── output workbook ────────────────────────────────────────────────────────────
print("Writing workbook …")
out = pd.DataFrame({
    "BP Number": df["BP Number"], "Cuscode": df["Cuscode"], "NAME1": df["NAME1"], "COUNTRY": df["COUNTRY"],
    "raw_STREET2": df["STREET2"], "raw_STREET3": df["STREET3"], "raw_STREET4": df["STREET4"],
    "raw_STREET5": df["STREET5"], "raw_Dist": df["Dist"], "raw_Other_city": df["Other city"],
    "raw_City": df["City"], "raw_POSTAL": df["POSTAL CODE"],
    "addr_house_no": res["house"], "addr_building": res["bld"], "addr_moo": res["moo"],
    "addr_soi": res["soi"], "addr_road": res["road"],
    "addr_subdistrict": res["sub"], "addr_district": res["dist"], "addr_province": res["prov"],
    "addr_postal": res["postal"], "addr_country": res["country"],
    "addr_subdistrict_en": res["sub_en"], "addr_district_en": res["dist_en"], "addr_province_en": res["prov_en"],
    "addr_tambon_id": res["tid"], "addr_district_id": res["did"], "addr_province_id": res["pid"],
    "addr_full_th": res["full_th"], "addr_full_en": res["full_en"],
    "addr_status": res["status"], "addr_confidence": res["conf"], "addr_lang": res["lang"],
    "addr_flags": res["flags"],
})
ADDED = [c for c in out.columns if c.startswith("addr_")]

wb = Workbook(); del wb["Sheet"]; ws = wb.create_sheet("Addresses")
hf = Font(bold=True, color="FFFFFF", name="Arial", size=10)
hfill = PatternFill("solid", start_color="37474F"); afill = PatternFill("solid", start_color="F57C00")
data = [[v if pd.notna(v) else None for v in row] for row in out.itertuples(index=False)]
rws = [list(out.columns)] + data
scol = list(out.columns).index("addr_status") + 1
fcol = list(out.columns).index("addr_full_th") + 1
aset = {list(out.columns).index(c) + 1 for c in ADDED}
for r_i, row in enumerate(rws, 1):
    for c_i, val in enumerate(row, 1):
        cell = ws.cell(r_i, c_i, val)
        if r_i == 1:
            cell.font = hf; cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.fill = afill if c_i in aset else hfill
        else:
            cell.font = Font(name="Arial", size=10)
            if c_i == scol and val in STATUS_FILL: cell.fill = STATUS_FILL[val]
            elif c_i == fcol and row[scol-1] in STATUS_FILL: cell.fill = STATUS_FILL[row[scol-1]]
ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
for c_i in range(1, len(rws[0]) + 1):
    w = max(len(str(r[c_i-1] or "")) for r in rws[:80])
    ws.column_dimensions[ws.cell(1, c_i).column_letter].width = min(w + 2, 48)

summ = Counter(res["status"])
ws2 = wb.create_sheet("Status_Summary")
ssum = [["addr_status","count","pct"]] + [[k, summ.get(k,0), f"{100*summ.get(k,0)/N:.1f}%"] for k in RANK]
for r_i, row in enumerate(ssum, 1):
    for c_i, val in enumerate(row, 1):
        cell = ws2.cell(r_i, c_i, val)
        if r_i == 1: cell.font = Font(bold=True)
        elif c_i == 1 and val in STATUS_FILL: cell.fill = STATUS_FILL[val]
for c_i in (1,2,3): ws2.column_dimensions[ws2.cell(1,c_i).column_letter].width = 18

# ── Notes / legend sheet ───────────────────────────────────────────────────────
ws3 = wb.create_sheet("Notes")
def put(r, txt, bold=False, fill=None, col=1):
    c = ws3.cell(r, col, txt)
    c.font = Font(bold=bold, name="Arial", size=12 if (bold and col==1 and r==1) else (11 if bold else 10),
                  color="FFFFFF" if fill else "000000")
    if fill: c.fill = fill
    c.alignment = Alignment(wrap_text=True, vertical="top")
    return c
r = 1
put(r, "TH44 Address Cleaning V2 — วิธีการและคำอธิบาย (Method & Legend)", True); r += 2
put(r, "วิธีการ (Method)", True); r += 1
for ln in [
    "• ใช้ฐานข้อมูลไปรษณีย์ไทยฉบับทางการ (thai-postal-codes_V2) เป็นตัวตัดสิน — ไม่เดา ไม่ใช้ LLM ไม่ต้องต่ออินเทอร์เน็ต",
    "• จังหวัด = lookup จาก POSTAL CODE โดยตรง (postal→จังหวัด 98.6% ออกค่าเดียว)",
    "• อำเภอ/ตำบล = ใช้ POSTAL ย่อรายการ candidate ก่อน แล้วจับคู่ข้อความที่กรอกด้วยโมเดล n-gram + cosine ที่เขียนเอง (ทำงานออฟไลน์)",
    "• ข้อมูลดิบ (คอลัมน์ raw_*) ไม่ถูกแก้ไขเลย — ค่าใหม่เขียนลงคอลัมน์ที่ถูกต้องของมัน รวมการย้ายข้อมูลที่กรอกผิดช่อง",
    "• ให้ที่อยู่ 2 ภาษา: ไทย (addr_full_th) + อังกฤษ (addr_full_en) จากชื่อทางการในฐานข้อมูล",
    "• addr_confidence = คะแนนความใกล้เคียงจริงของการจับคู่ (0–100)",
]: put(r, ln); r += 1
r += 1
put(r, "สถานะ (addr_status) และสี", True); r += 1
for st, desc in [
    ("VERIFIED", "ตรงกับชื่อทางการในฐานข้อมูลเป๊ะ ไม่ต้องแก้"),
    ("AUTO_FIXED", "แก้อัตโนมัติแล้ว (ขยายตัวย่อ / เติม prefix / เติมจาก postal / ย้ายช่อง) และตรงกับฐานข้อมูลทางการ"),
    ("FUZZY_FIXED", "แก้คำพิมพ์ผิด/romanization ด้วยการจับคู่ความใกล้เคียง — แนะนำให้ตรวจทาน"),
    ("NEEDS_REVIEW", "ต้องตรวจ: postal ไม่อยู่ในฐานข้อมูล/ไม่มี postal, จังหวัดขัดกับ postal, หรือทั้งตำบลและอำเภอไม่ตรงกับ postal"),
    ("FOREIGN", "ที่อยู่ต่างประเทศ — ประกอบตามต้นฉบับ"),
]:
    put(r, st, fill=STATUS_FILL[st]); ws3.cell(r, 2, desc).font = Font(name="Arial", size=10)
    ws3.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top"); r += 1
r += 1
put(r, "ความหมาย addr_flags ที่พบบ่อย", True); r += 1
for ln in [
    "province:standardized — ปรับชื่อจังหวัดให้เป็นรูปแบบทางการ",
    "province:conflict — จังหวัดที่กรอก (City) ขัดกับจังหวัดที่ได้จาก postal → ยึด postal เป็นหลัก โปรดตรวจ",
    "subdistrict:fuzzy / district:fuzzy — จับคู่แบบใกล้เคียง (พิมพ์ผิด/romanization)",
    "subdistrict:unmatched — ตำบลที่กรอกไม่ตรงกับ candidate ของ postal (romanization ยาก หรือ postal/ตำบลขัดกัน) — อำเภอ+จังหวัดยังถูก",
    "subdistrict:moved_from(ช่อง) — พบชื่อตำบลในช่องอื่น (กรอกผิดช่อง) ระบบย้ายให้ถูก",
    "district:from_postal(single) — postal นี้มีอำเภอเดียว เติมให้อัตโนมัติ",
    "wrong_column:STREET2_is_name — STREET2 เป็นชื่ออาคาร/สถานที่ ไม่ใช่เลขที่ → ย้ายไป addr_building",
    "road:abbrev_expanded — ขยาย ถ./ซ. เป็น ถนน/ซอย",
    "postal_not_in_db / no_postal — postal ไม่อยู่ในฐานข้อมูล หรือไม่มี postal → ใช้ fallback จับคู่ทั้งฐานข้อมูล + ตั้งให้ตรวจ",
]: put(r, ln); r += 1
r += 1
put(r, "โครงสร้างคอลัมน์ (Schema)", True); r += 1
for ln in [
    "BP Number, Cuscode, NAME1, COUNTRY — คอลัมน์อ้างอิง",
    "raw_* — ข้อมูลดิบเดิม (ไม่แก้ไข)",
    "addr_house_no, addr_building, addr_moo, addr_soi, addr_road — รายละเอียดที่อยู่ (แยกส่วน)",
    "addr_subdistrict/district/province (ไทย) + _en (อังกฤษ) — ชื่อทางการจากฐานข้อมูล",
    "addr_tambon_id / district_id / province_id — รหัสทางการ (ไว้ join ข้อมูลในอนาคต)",
    "addr_full_th / addr_full_en — ที่อยู่ประกอบเต็ม 2 ภาษา",
    "addr_status / addr_confidence / addr_lang / addr_flags — สถานะ / ความมั่นใจ / ภาษา / บันทึกการแก้",
]: put(r, ln); r += 1
r += 1
put(r, "ข้อจำกัด (Limitations)", True); r += 1
for ln in [
    "• ฐานข้อมูลครอบคลุม ~99.6% ของ postal ที่พบ; ส่วนที่เหลือใช้ fallback + flag (ดู postal_not_in_db/no_postal)",
    "• แถวที่ postal ขัดกับอำเภอ/จังหวัดที่กรอก จะถูกตั้ง NEEDS_REVIEW ให้ตรวจ (มักเป็นข้อมูลผิดจริง)",
    "• ตำบลที่เป็น romanization แปลกบางตัวอาจจับคู่ไม่ได้ (flag ไว้) แต่อำเภอ+จังหวัดยังถูกต้อง",
]: put(r, ln); r += 1
ws3.column_dimensions["A"].width = 30; ws3.column_dimensions["B"].width = 95

wb.save(OUT)

print("\n" + "="*60 + "\n  V2 ADDRESS CLEANING COMPLETE")
print(f"  Rows: {N:,}")
for k in RANK:
    if summ.get(k): print(f"    {k:13} {summ[k]:>6,}  ({100*summ[k]/N:4.1f}%)")
print(f"  Output: {OUT.name}\n" + "="*60)
for st in ["VERIFIED","AUTO_FIXED","FUZZY_FIXED","NEEDS_REVIEW"]:
    ex = res[res["status"] == st].head(1)
    if len(ex): print(f"  [{st}] {ex.iloc[0]['full_th'][:88]}")