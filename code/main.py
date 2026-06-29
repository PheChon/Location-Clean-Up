"""
TH44 Customer Master — Audit-Driven Data Cleaning Pipeline
===========================================================
Non-destructive: raw values untouched; adds _std / _valid companion cols.
Output: 3-sheet .xlsx  (Cleaned | Issues | Notes)
"""

import re
import warnings
from datetime import date as _date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings("ignore")

# ── CONFIG ─────────────────────────────────────────────────────────────────────
INPUT_PATH  = Path("/mnt/user-data/uploads/24062026_TH44_all_CUSTOMERS_excl_flag_del.xlsx")
OUTPUT_PATH = Path("/mnt/user-data/outputs/TH44_Customers_Cleaned.xlsx")
TODAY       = pd.Timestamp(_date.today())
MIN_DATE    = pd.Timestamp("2006-01-01")

DATE_COLS  = ["CHANGE DATE", "Create Date Comp", "Create Date Sale", "Create Date Gen"]
PHONE_COLS = ["TELEPHONE1", "Telephone 2", "Mobile no", "Mobile no 2"]
EMAIL_COLS = [
    "E-Mail Address1", "E-Mail Address2", "E-Mail Address3",
    "E-Mail Address4", "E-Mail Address5",
]
EMPTY_COLS    = ["Deletion Flag", "Cust_group3", "Customer group"]
CONSTANT_COLS = [
    "Company", "Saleorg", "Channel", "Division", "CCA",
    "Cust_group4", "EB360 RP/TPI Reference Num", "Cust. E-TAX",
]
REDUNDANT_COUNTRY_COL = "Country"   # col-68 duplicate of COUNTRY (col-62)

# TINs appearing 80+ times confirmed as shared / placeholder from audit
PLACEHOLDER_TINS = {
    "9999999999902",   # all-9s placeholder
    "0994000158378",   # 228× — shared/HQ TIN
    "0994000159382",   # 183× — shared/HQ TIN
    "0994000423179",   #  81× — shared/HQ TIN
}

CATEGORICAL_ALLOWED: dict[str, set] = {
    "Shp. Cond.":        {"01"},
    "Group":             {"TH10", "TH11"},
    "Acct assgmt Group": {"01", "02", "06"},
    "TaxC":              {"1", "4"},
}

_issues: list[dict] = []
_notes:  list[dict] = []

# ── HELPERS ────────────────────────────────────────────────────────────────────
def flag(bp: str, field: str, reason: str, raw) -> None:
    _issues.append({
        "BP Number": bp,
        "Field":     field,
        "Reason":    reason,
        "Raw Value": "" if pd.isna(raw) else str(raw),
    })

def note(section: str, detail: str) -> None:
    _notes.append({"Section": section, "Detail": detail})

def digits_only(val) -> str:
    return re.sub(r"\D", "", str(val)) if pd.notna(val) else ""

def safe_name(col: str) -> str:
    """Column name → safe identifier for companion column naming."""
    return re.sub(r"[^A-Za-z0-9]", "_", col)

def _prep(frame: pd.DataFrame) -> pd.DataFrame:
    """Convert pd.NA → None and Timestamps → date for openpyxl."""
    out = frame.copy()
    for c in out.columns:
        if pd.api.types.is_datetime64_any_dtype(out[c]):
            out[c] = [v.date() if pd.notna(v) else None for v in out[c]]
        else:
            out[c] = [v if pd.notna(v) else None for v in out[c]]
    return out

FILLS = {
    "Cleaned": PatternFill("solid", start_color="1A73E8"),
    "Issues":  PatternFill("solid", start_color="EA4335"),
    "Notes":   PatternFill("solid", start_color="34A853"),
}
# Companion / added columns  — header: orange  |  data: light amber
ADDED_HDR_FILL  = PatternFill("solid", start_color="F57C00")  # deep orange
ADDED_DATA_FILL = PatternFill("solid", start_color="FFF3E0")  # light amber

def write_sheet(wb: Workbook, name: str, data: pd.DataFrame, added_cols=None) -> None:
    ws       = wb.create_sheet(title=name)
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    dat_font = Font(name="Arial", size=10)
    hdr_fill = FILLS[name]

    # Build 1-based index set of companion / added columns for this sheet
    added_idxs = {
        i for i, c in enumerate(data.columns, 1)
        if added_cols and c in added_cols
    }

    rows = list(dataframe_to_rows(_prep(data), index=False, header=True))
    for r_i, row in enumerate(rows, 1):
        for c_i, val in enumerate(row, 1):
            cell     = ws.cell(row=r_i, column=c_i, value=val)
            is_added = c_i in added_idxs
            if r_i == 1:
                cell.font      = hdr_font
                cell.fill      = ADDED_HDR_FILL if is_added else hdr_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            else:
                cell.font = dat_font
                if is_added:
                    cell.fill = ADDED_DATA_FILL
                if isinstance(val, _date):
                    cell.number_format = "YYYY-MM-DD"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Width from header + first 50 data rows (fast, avoids full scan)
    sample = rows[: min(51, len(rows))]
    for c_i in range(1, len(rows[0]) + 1):
        max_len = max(len(str(r[c_i - 1] or "")) for r in sample)
        ws.column_dimensions[ws.cell(1, c_i).column_letter].width = min(max_len + 2, 55)

# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — LOAD
# ══════════════════════════════════════════════════════════════════════════════
print("─" * 62)
print("STEP 1  Loading …")
OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
df             = pd.read_excel(INPUT_PATH, dtype=str)
ORIG_ROWS, ORIG_COLS = df.shape
print(f"        {ORIG_ROWS:,} rows × {ORIG_COLS} cols")
assert df["BP Number"].nunique() == ORIG_ROWS, "BP Number not unique on load — abort!"
note("Load", f"Source: {INPUT_PATH.name}  |  {ORIG_ROWS:,} rows × {ORIG_COLS} cols")

# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — STRUCTURAL CLEANUP
# ══════════════════════════════════════════════════════════════════════════════
print("STEP 2  Structural cleanup …")

# 2a. Log constant columns (before touching anything)
for c in CONSTANT_COLS:
    if c in df.columns:
        val = df[c].dropna().iloc[0] if df[c].notna().any() else "N/A"
        note("Constant col (kept)", f"{c!r} = {val!r}")

# 2b. Drop fully-empty columns
drop_empty = [c for c in EMPTY_COLS if c in df.columns]
df.drop(columns=drop_empty, inplace=True)
for c in drop_empty:
    note("Empty col (dropped)", c)

# 2c. Coalesce redundant Country (col-68) into COUNTRY (col-62), then drop
if REDUNDANT_COUNTRY_COL in df.columns:
    fill_mask = df["COUNTRY"].isna() & df[REDUNDANT_COUNTRY_COL].notna()
    n_filled  = int(fill_mask.sum())
    df.loc[fill_mask, "COUNTRY"] = df.loc[fill_mask, REDUNDANT_COUNTRY_COL]
    df.drop(columns=[REDUNDANT_COUNTRY_COL], inplace=True)
    note("Redundant col (dropped)",
         f"'{REDUNDANT_COUNTRY_COL}' coalesced into COUNTRY ({n_filled} fills)")

print(f"        Dropped {len(drop_empty)} empty + 1 redundant → {df.shape[1]} cols remain")

# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — TEXT & FLAG HYGIENE  (column-aware — never blindly strip codes)
# ══════════════════════════════════════════════════════════════════════════════
print("STEP 3  Text hygiene & flag normalisation …")

# 3a. Trim whitespace on every text column; collapse multi-spaces; blank → NA
_DATE_SET = set(DATE_COLS)
for c in df.columns:
    if c in _DATE_SET:
        continue
    df[c] = (
        df[c].astype(str)
             .str.strip()
             .str.replace(r"\s{2,}", " ", regex=True)
             .replace({"": pd.NA, "nan": pd.NA, "None": pd.NA, "NaN": pd.NA, "<NA>": pd.NA})
    )

# 3b. Blocked: SAP flag — lowercase 'x' is valid but non-standard → uppercase
if "Blocked" in df.columns:
    df["Blocked"] = df["Blocked"].str.upper()

# 3c. Categorical anomaly detection (flag unexpected values, keep them)
cat_issue_count = 0
for col, allowed in CATEGORICAL_ALLOWED.items():
    if col not in df.columns:
        continue
    mask = df[col].notna() & ~df[col].isin(allowed)
    for _, row in df[mask].iterrows():
        flag(row["BP Number"], col,
             f"Unexpected value; allowed = {sorted(allowed)}", row[col])
        cat_issue_count += 1

print(f"        Whitespace trimmed, Blocked normalised, "
      f"{cat_issue_count} categorical anomalies flagged")

# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — DATE COLUMNS
# ══════════════════════════════════════════════════════════════════════════════
print("STEP 4  Parsing date columns …")
for c in DATE_COLS:
    if c not in df.columns:
        continue
    parsed = pd.to_datetime(df[c], errors="coerce").dt.normalize()
    df[c]  = parsed
    oob    = df[parsed.notna() & ((parsed < MIN_DATE) | (parsed > TODAY))]
    for _, row in oob.iterrows():
        flag(row["BP Number"], c,
             f"Date outside valid range [{MIN_DATE.date()} – {TODAY.date()}]", row[c])
    print(f"        {c}: {parsed.notna().sum():,} valid | "
          f"{parsed.isna().sum()} null | {len(oob)} OOB")

# ══════════════════════════════════════════════════════════════════════════════
# STEP 5a — TAX NUMBER3  (Thai TIN = exactly 13 digits)
# ══════════════════════════════════════════════════════════════════════════════
print("STEP 5a TIN validation (Tax Number3) …")
if "Tax Number3" in df.columns:
    df["Tax_Number3_digits"] = df["Tax Number3"].apply(digits_only)

    def _tin_valid(row: pd.Series) -> str:
        raw  = row["Tax Number3"]
        digs = row["Tax_Number3_digits"]
        bp   = row["BP Number"]
        if pd.isna(raw):
            return "MISSING"
        if len(digs) != 13:
            flag(bp, "Tax Number3",
                 f"Wrong digit count ({len(digs)}, expected 13)", raw)
            return "FLAG_LENGTH"
        if digs in PLACEHOLDER_TINS or set(digs) == {"9"}:
            flag(bp, "Tax Number3",
                 "Shared or placeholder TIN (informational — verify before any merge)", raw)
            return "FLAG_PLACEHOLDER"
        return "OK"

    df["Tax_Number3_valid"] = df.apply(_tin_valid, axis=1)
    print(f"        {df['Tax_Number3_valid'].value_counts().to_dict()}")

# ══════════════════════════════════════════════════════════════════════════════
# STEP 5b — POSTAL CODE  (country-aware: Thai = 5 digits, foreign = 4-10)
# ══════════════════════════════════════════════════════════════════════════════
print("STEP 5b Postal code validation (country-aware) …")
if "POSTAL CODE" in df.columns:
    def _postal_valid(row: pd.Series) -> str:
        raw     = row["POSTAL CODE"]
        bp      = row["BP Number"]
        cval    = row["COUNTRY"]
        country = str(cval).strip() if pd.notna(cval) else "Thailand"
        if pd.isna(raw):
            return "MISSING"
        digs    = re.sub(r"\D", "", str(raw))
        is_thai = country in ("Thailand", "TH", "")
        if is_thai:
            if re.fullmatch(r"\d{5}", digs):
                return "OK"
            flag(bp, "POSTAL CODE",
                 f"Thai postal must be exactly 5 digits (got: {digs!r})", raw)
            return "FLAG"
        else:
            if 4 <= len(digs) <= 10:
                return "OK_FOREIGN"
            flag(bp, "POSTAL CODE",
                 f"Foreign postal suspicious digit count ({len(digs)})", raw)
            return "FLAG"

    df["Postal_Code_valid"] = df.apply(_postal_valid, axis=1)
    print(f"        {df['Postal_Code_valid'].value_counts().to_dict()}")

# ══════════════════════════════════════════════════════════════════════════════
# STEP 5c — PHONES (x4 columns)
# Logic: Thai local 0XX → +66XX | 00XX → +XX | +XX → keep | 66XX → +66XX
# ══════════════════════════════════════════════════════════════════════════════
print("STEP 5c Phone normalisation …")

def _phone_std_valid(raw, bp: str, field: str) -> tuple:
    if pd.isna(raw) or str(raw).strip() == "":
        return None, "MISSING"
    s        = str(raw).strip()
    has_plus = s.startswith("+")
    digs     = re.sub(r"\D", "", s)

    if len(digs) < 7:
        flag(bp, field, f"Too short ({len(digs)} digits) to be a valid phone", raw)
        return s, "FLAG_SHORT"
    if len(digs) > 15:                             # ITU E.164 max = 15 digits
        flag(bp, field,
             f"Digit count ({len(digs)}) > 15 — may be range/extension combined", raw)
        return s, "FLAG_LONG"

    if digs.startswith("0") and 9 <= len(digs) <= 10:   # Thai local: 02-XXX / 08X-XXX
        return "+66" + digs[1:], "OK"
    if digs.startswith("00") and 10 <= len(digs) <= 14: # 00-prefix international
        return "+" + digs[2:], "OK"
    if has_plus and 10 <= len(digs) <= 15:               # already E.164 (+XX…)
        return "+" + digs, "OK"
    if digs.startswith("66") and 10 <= len(digs) <= 12: # Thai country code without +
        return "+" + digs, "OK"

    flag(bp, field, "Unrecognised phone format — check manually", raw)
    return ("+" if has_plus else "") + digs, "FLAG_FORMAT"

for col in PHONE_COLS:
    if col not in df.columns:
        continue
    base    = safe_name(col)
    results = df.apply(
        lambda r, c=col: _phone_std_valid(r[c], r["BP Number"], c), axis=1
    )
    df[f"{base}_std"]   = [r[0] for r in results]
    df[f"{base}_valid"] = [r[1] for r in results]
    print(f"        {col}: {df[f'{base}_valid'].value_counts().to_dict()}")

# ══════════════════════════════════════════════════════════════════════════════
# STEP 5d — EMAILS (x5 columns)
# Auto-fix: strip spaces, -com → .com | Flag unfixable
# ══════════════════════════════════════════════════════════════════════════════
print("STEP 5d E-mail validation …")
_EMAIL_RE   = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]{2,}$")
_EMAIL_FIXES = [("-com", ".com"), ("_com", ".com"), (",.com", ".com")]

def _email_std_valid(raw, bp: str, field: str) -> tuple:
    if pd.isna(raw) or str(raw).strip() == "":
        return None, "MISSING"
    orig = str(raw).strip()
    std  = orig.replace(" ", "")             # strip embedded spaces
    for bad, good in _EMAIL_FIXES:
        std = std.replace(bad, good)
    if _EMAIL_RE.match(std):
        return std, "OK" if std == orig else "OK_FIXED"
    flag(bp, field, "Invalid e-mail format (could not auto-fix)", orig)
    return std, "FLAG"

for col in EMAIL_COLS:
    if col not in df.columns:
        continue
    base    = safe_name(col)
    results = df.apply(
        lambda r, c=col: _email_std_valid(r[c], r["BP Number"], c), axis=1
    )
    df[f"{base}_std"]   = [r[0] for r in results]
    df[f"{base}_valid"] = [r[1] for r in results]
    print(f"        {col}: {df[f'{base}_valid'].value_counts(dropna=False).to_dict()}")

# ══════════════════════════════════════════════════════════════════════════════
# STEP 6 — DUPLICATE DETECTION  (FLAG ONLY — never merge, never delete)
# Signature: normalised name + postal + phone digits (only among rows with phone)
# ══════════════════════════════════════════════════════════════════════════════
print("STEP 6  Duplicate candidate detection …")
df["_sig"] = (
    df["NAME1"].fillna("").str.lower().str.strip()
    + "|" + df["POSTAL CODE"].fillna("")
    + "|" + df["TELEPHONE1"].apply(digits_only)
)
with_phone = df["TELEPHONE1"].notna()
sig_counts = df.loc[with_phone, "_sig"].value_counts()
dup_sigs   = set(sig_counts[sig_counts > 1].index)

df["duplicate_candidate"] = df.apply(
    lambda r: "FLAG"
    if (pd.notna(r["TELEPHONE1"]) and r["_sig"] in dup_sigs)
    else "",
    axis=1,
)
n_dup = (df["duplicate_candidate"] == "FLAG").sum()
for sig, cnt in sig_counts[sig_counts > 1].items():
    bp_list = " | ".join(df.loc[df["_sig"] == sig, "BP Number"].tolist())
    flag(bp_list, "NAME1 + POSTAL CODE + TELEPHONE1",
         f"Possible duplicate group ({cnt} records — review before any merge)", sig)

df.drop(columns=["_sig"], inplace=True)
print(f"        {n_dup} rows flagged as duplicate_candidate")

# ══════════════════════════════════════════════════════════════════════════════
# STEP 7 — ASSEMBLE OUTPUT WORKBOOK  (Cleaned | Issues | Notes)
# ══════════════════════════════════════════════════════════════════════════════
print("STEP 7  Assembling output workbook …")

issues_df = pd.DataFrame(_issues, columns=["BP Number", "Field", "Reason", "Raw Value"])

note("─" * 40, "")
note("Stats", f"Original:             {ORIG_ROWS:,} rows × {ORIG_COLS} cols")
note("Stats", f"Cleaned:              {len(df):,} rows × {df.shape[1]} cols")
note("Stats", f"Total issues flagged: {len(issues_df):,}")
note("Stats", f"Duplicate candidates: {n_dup:,}")
note("─" * 40, "")
note("Issues breakdown by field", "")
for field, cnt in issues_df["Field"].value_counts().head(20).items():
    note("  →", f"{field}: {cnt:,}")
note("─" * 40, "")
note("Companion columns added", "")
for c in df.columns:
    if (c.endswith("_std") or c.endswith("_valid")
            or c in {"Tax_Number3_digits", "duplicate_candidate"}):
        note("  +", c)

notes_df = pd.DataFrame(_notes, columns=["Section", "Detail"])

wb = Workbook()
del wb["Sheet"]
# Identify every companion column added by the pipeline
added_col_names = {
    c for c in df.columns
    if c.endswith("_std") or c.endswith("_valid")
    or c in {"Tax_Number3_digits", "duplicate_candidate"}
}
write_sheet(wb, "Cleaned", df,        added_col_names)  # highlight companion cols
write_sheet(wb, "Issues",  issues_df)
write_sheet(wb, "Notes",   notes_df)
wb.save(OUTPUT_PATH)
print(f"        Saved → {OUTPUT_PATH}")

# ══════════════════════════════════════════════════════════════════════════════
# STEP 8 — VALIDATION GATE  (evidence before claims)
# ══════════════════════════════════════════════════════════════════════════════
print("STEP 8  Validation gate …")
fails: list[str] = []

# Row / key integrity
if len(df) != ORIG_ROWS:
    fails.append(f"Row count drifted: {ORIG_ROWS} → {len(df)}")
if df["BP Number"].nunique() != len(df):
    fails.append("BP Number uniqueness broken after cleaning")

# Blocked flag sanity
if "Blocked" in df.columns:
    bad_blocked = set(df["Blocked"].dropna().unique()) - {"X"}
    if bad_blocked:
        fails.append(f"Blocked still contains non-X values: {bad_blocked}")

# Date columns properly typed
for c in DATE_COLS:
    if c in df.columns and not pd.api.types.is_datetime64_any_dtype(df[c]):
        fails.append(f"Date column not typed as datetime: {c!r}")

# All companion columns present
expected = (
    [f"{safe_name(c)}_valid" for c in PHONE_COLS]
    + [f"{safe_name(c)}_valid" for c in EMAIL_COLS]
    + ["Tax_Number3_valid", "Postal_Code_valid"]
)
for vcol in expected:
    if vcol not in df.columns:
        fails.append(f"Missing companion column: {vcol!r}")

# Issues table populated (audit guarantees known issues)
if len(issues_df) == 0:
    fails.append("Issues table is empty — expected known flagged records")

# No remaining leading/trailing whitespace in text columns
padded_count = sum(
    int((df[c].dropna().astype(str) != df[c].dropna().astype(str).str.strip()).sum())
    for c in df.select_dtypes("object").columns
)
if padded_count:
    fails.append(f"{padded_count} cells still have leading/trailing whitespace")

for msg in fails:
    print(f"  ✗  {msg}")
if not fails:
    print("  ✓  All validation checks passed")

# ── FINAL SUMMARY ──────────────────────────────────────────────────────────────
tin_flags    = (df["Tax_Number3_valid"].str.startswith("FLAG")).sum() \
               if "Tax_Number3_valid" in df.columns else 0
postal_flags = (df["Postal_Code_valid"].str.startswith("FLAG")).sum() \
               if "Postal_Code_valid" in df.columns else 0

print()
print("═" * 62)
print("  TH44 Customer Master — Pipeline Complete")
print(f"  Rows:                 {ORIG_ROWS:,} → {len(df):,}  (unchanged)")
print(f"  Cols:                 {ORIG_COLS}  → {df.shape[1]}  "
      f"(+{df.shape[1] - ORIG_COLS} companion cols)")
print(f"  Issues logged:        {len(issues_df):,}")
print(f"    ∟ TIN flags:        {tin_flags:,}")
print(f"    ∟ Postal flags:     {postal_flags:,}")
print(f"    ∟ Dup candidates:   {n_dup:,}")
print(f"  Validation:           {'PASSED ✓' if not fails else 'FAILED ✗'}")
print(f"  Output:               {OUTPUT_PATH.name}")
print("═" * 62)